# Copyright (C) [2025] [阿炜]
# 本程序是自由软件：你可以根据 GNU GPL v3 协议重新分发或修改它。

DISCLAIMER_TEXT = """
开源说明

本软件基于 PyQt5 开发，遵循 GPL v3 开源协议。
您有权获得、修改和重新分发本软件的源代码。
源代码获取地址：https://github.com/WeijieG/ReagentManagementSystem.git
如无法使用、需求变更或者需要技术支持，请前往GitHub项目仓库提交Issues
详细说明请前往 更多-关于-开源说明 查看

用前须知

本软件为方便库存管理而开发，仅供学习参考，禁止使用本软件进行商业行为。用户在使用本软件前应确保：

1. 定期备份重要数据！！！！！数据无价！请务必定期备份！！！！！
2. 符合所有适用的法律法规和行业标准
3. 遵循实验室安全操作规程
4. 禁止使用本软件进行任何商业行为
5. 软件运行前提需要保证应用完整性

使用本软件即表示您已阅读、理解并同意接受上述条款的全部内容。如有异议，请停止使用本软件。

"""

#                            _ooOoo_
#                           o8888888o
#                           88" . "88
#                           (| -_- |)
#                            O\ = /O
#                        ____/`---'\____
#                      .   ' \\| |// `.
#                       / \\||| : |||// \
#                     / _||||| -:- |||||- \
#                       | | \\\ - /// | |
#                     | \_| ''\---/'' | |
#                      \ .-\__ `-` ___/-. /
#                   ___`. .' /--.--\ `. . __
#                ."" '< `.___\_<|>_/___.' >'"".
#               | | : `- \`.;`\ _ /`;.`/ - ` : | |
#                 \ \ `-. \_ __\ /__ _/ .-` / /
#         ======`-.____`-.___\_____/___.-`____.-'======
#                            `=---='
#
#         .............................................
#                  佛祖保佑             永无BUG

import sys
import sqlite3
import shutil
import os
import csv
import configparser
import requests
import subprocess
import tempfile
import hashlib
from datetime import datetime
from PyQt5.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout, 
                             QTableView, QPushButton, QComboBox, QLineEdit, QDateEdit, 
                             QLabel, QMessageBox, QHeaderView, QFormLayout, QDialog, 
                             QDialogButtonBox, QGroupBox, QStackedWidget, QStatusBar, 
                             QTabWidget, QAbstractItemView, QFileDialog, QInputDialog, QFrame, QCheckBox,
                             QListWidget, QListWidgetItem,QCompleter,QSizePolicy, QProgressBar, QTextEdit)
from PyQt5.QtGui import QStandardItemModel, QStandardItem, QIntValidator, QBrush, QColor, QFont,QIcon
from PyQt5.QtCore import Qt, QDate, QSortFilterProxyModel, pyqtSignal, QDateTime, QThread, QTimer
# 使用openpyxl创建Excel文件
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Side, Font
from openpyxl.utils import get_column_letter
import json
import re


# 确定图标文件路径
if getattr(sys, 'frozen', False):
    # 打包后的可执行文件
    application_path = os.path.dirname(sys.executable)
    ICON_PATH = os.path.join(application_path, "reagent.ico")
else:
    # 开发环境
    application_path = os.path.dirname(os.path.abspath(__file__))
    ICON_PATH = os.path.join(application_path, "reagent.ico")

ICON_PATH = os.path.join(application_path, "reagent.ico")

# 设置应用图标（如果存在）
if os.path.exists(ICON_PATH):
    try:
        from PyQt5.QtGui import QIcon
        app = QApplication.instance()
        app.setWindowIcon(QIcon(ICON_PATH))
    except:
        pass


# 确定数据库文件路径
if getattr(sys, 'frozen', False):
    # 打包后的可执行文件 - 使用AppData目录
    appdata_path = os.getenv('APPDATA')
    if appdata_path:
        DB_DIR = os.path.join(appdata_path, "ReagentManagementSystem")
        # 确保目录存在
        os.makedirs(DB_DIR, exist_ok=True)
        DB_FILE = os.path.join(DB_DIR, "ReagentWarehouseData.db")
    else:
        # 备用方案：使用可执行文件目录
        DB_DIR = os.path.dirname(sys.executable)
        DB_FILE = os.path.join(DB_DIR, "ReagentWarehouseData.db")
else:
    # 开发环境
    DB_DIR = os.path.dirname(os.path.abspath(__file__))
    DB_FILE = os.path.join(DB_DIR, "ReagentWarehouseData.db")

def get_image_path(file_name):
    """获取图片的绝对路径，并检查文件是否存在"""
    # 构建图片路径
    image_path = os.path.join(application_path, "images", file_name)
    
    # 检查文件是否存在
    if not os.path.exists(image_path):
        print(f"警告: 图片文件不存在 - {image_path}")
        return ""  # 返回空字符串而不是引发错误
    
    # 规范化路径（适用于Windows）
    return os.path.normpath(image_path).replace("\\", "/")
    
class ReagentDatabase:
    def __init__(self):
        print(f"Database path: {DB_FILE}")  # 调试输出
        # 确保目录存在
        db_dir = os.path.dirname(DB_FILE)
        if not os.path.exists(db_dir):
            os.makedirs(db_dir, exist_ok=True)
        
        try:
            self.conn = sqlite3.connect(DB_FILE)
            self.cursor = self.conn.cursor()
            self.create_tables()
        except sqlite3.Error as e:
            QMessageBox.critical(None, "数据库错误", f"无法打开数据库文件:\n{DB_FILE}\n错误: {str(e)}")
            sys.exit(1)
    
    def create_tables(self):
        # 创建试剂名称表
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS reagent_names (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                gtin TEXT UNIQUE
            )
        ''')
        
        # 创建试剂信息表
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS reagents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL,
                batch TEXT NOT NULL,
                production_date TEXT NOT NULL,
                expiry_date TEXT NOT NULL,
                quantity INTEGER NOT NULL,
                gtin TEXT,  -- 新增GTIN字段
                qrcode TEXT,  -- 新增二维码原始数据字段
                FOREIGN KEY (name) REFERENCES reagent_names(name) ON DELETE CASCADE,
                UNIQUE (name, batch)  -- 添加复合唯一约束
            )
        ''')
        
        # 创建入库记录表
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS inbound (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                reagent_id INTEGER NOT NULL,
                reagent_name TEXT NOT NULL,
                reagent_batch TEXT NOT NULL,
                date TEXT NOT NULL,
                quantity INTEGER NOT NULL,
                operator TEXT,
                gtin TEXT,  -- 新增GTIN字段
                remaining_quantity INTEGER NOT NULL,  -- 新增：入库后剩余库存
                FOREIGN KEY (reagent_id) REFERENCES reagents (id)
            )
        ''')
        
        # 创建出库记录表
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS outbound (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                reagent_id INTEGER NOT NULL,
                reagent_name TEXT NOT NULL,
                reagent_batch TEXT NOT NULL,
                date TEXT NOT NULL,
                quantity INTEGER NOT NULL,
                operator TEXT,
                gtin TEXT,  -- 新增GTIN字段
                remaining_quantity INTEGER NOT NULL,  -- 新增：出库后剩余库存
                FOREIGN KEY (reagent_id) REFERENCES reagents (id)
            )
        ''')
        self.conn.commit()
    
    def get_reagents(self):
        self.cursor.execute('SELECT id, name, batch, production_date, expiry_date, quantity FROM reagents')
        return self.cursor.fetchall()
    
    def get_reagent_names(self):
        self.cursor.execute('SELECT name FROM reagent_names ORDER BY name')
        return [item[0] for item in self.cursor.fetchall()]
    
    def get_all_reagent_names(self):
        return self.get_reagent_names()
    
    def add_reagent_name(self, name, gtin=None):
        try:
            self.cursor.execute('''
                INSERT INTO reagent_names (name, gtin) VALUES (?, ?)
            ''', (name, gtin))
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False
    
    def get_name_by_gtin(self, gtin):
        self.cursor.execute('''
            SELECT name 
            FROM reagent_names 
            WHERE gtin = ?
        ''', (gtin,))
        result = self.cursor.fetchone()
        return result[0] if result else None

    def delete_reagent_name(self, name):
        # 检查是否有库存记录
        self.cursor.execute('SELECT COUNT(*) FROM reagents WHERE name = ?', (name,))
        count = self.cursor.fetchone()[0]
        
        if count > 0:
            return False, "该试剂名称有库存记录，无法删除"
        
        # 检查是否有出入库记录
        self.cursor.execute('SELECT COUNT(*) FROM inbound WHERE reagent_name = ?', (name,))
        count += self.cursor.fetchone()[0]
        
        self.cursor.execute('SELECT COUNT(*) FROM outbound WHERE reagent_name = ?', (name,))
        count += self.cursor.fetchone()[0]
        
        if count > 0:
            return False, "该试剂名称有出入库记录，无法删除"
        
        # 删除试剂名称
        self.cursor.execute('DELETE FROM reagent_names WHERE name = ?', (name,))
        self.conn.commit()
        return True, "删除成功"
    
    def get_batches_by_name(self, name):
        self.cursor.execute('SELECT batch FROM reagents WHERE name = ? ORDER BY batch', (name,))
        return [item[0] for item in self.cursor.fetchall()]
    
    def get_reagent_details(self, name, batch):
        self.cursor.execute('''
            SELECT id, production_date, expiry_date, quantity 
            FROM reagents 
            WHERE name = ? AND batch = ?
        ''', (name, batch))
        return self.cursor.fetchone()
    
    def get_name_by_batch(self, batch):
        self.cursor.execute('''
            SELECT name 
            FROM reagents 
            WHERE batch = ?
        ''', (batch,))
        result = self.cursor.fetchone()
        return result[0] if result else None
    
    def get_inbound_records(self, name_filter="", batch_filter="", start_date=None, end_date=None):
        query = '''
            SELECT id, reagent_name, reagent_batch, date, quantity, operator, remaining_quantity
            FROM inbound
            WHERE 1=1
        '''
        params = []
        
        if name_filter:
            query += " AND reagent_name LIKE ?"
            params.append(f"%{name_filter}%")
        
        if batch_filter:
            query += " AND reagent_batch LIKE ?"
            params.append(f"%{batch_filter}%")
        
        # 添加日期范围筛选
        if start_date:
            query += " AND date >= ?"
            params.append(start_date)
        
        if end_date:
            query += " AND date <= ?"
            params.append(end_date)
        
        self.cursor.execute(query, params)
        return self.cursor.fetchall()
    
    def get_outbound_records(self, name_filter="", batch_filter="", start_date=None, end_date=None):
        query = '''
            SELECT id, reagent_name, reagent_batch, date, quantity, operator, remaining_quantity
            FROM outbound
            WHERE 1=1
        '''
        params = []
        
        if name_filter:
            query += " AND reagent_name LIKE ?"
            params.append(f"%{name_filter}%")
        
        if batch_filter:
            query += " AND reagent_batch LIKE ?"
            params.append(f"%{batch_filter}%")
        
        # 添加日期范围筛选
        if start_date:
            query += " AND date >= ?"
            params.append(start_date)
        
        if end_date:
            query += " AND date <= ?"
            params.append(end_date)
        
        self.cursor.execute(query, params)
        return self.cursor.fetchall()
    
    def get_matching_reagent_names(self, prefix):
        self.cursor.execute('''
            SELECT name 
            FROM reagent_names 
            WHERE name LIKE ? 
            ORDER BY name
        ''', (f"{prefix}%",))
        return [item[0] for item in self.cursor.fetchall()]
    
    def get_matching_batches(self, prefix):
        self.cursor.execute('''
            SELECT DISTINCT batch 
            FROM reagents 
            WHERE batch LIKE ? 
            ORDER BY batch
        ''', (f"{prefix}%",))
        return [item[0] for item in self.cursor.fetchall()]
    
    def add_inbound(self, reagent_id, name, batch, date, quantity, operator, gtin=None):
        # 更新库存前先查询当前库存
        self.cursor.execute('SELECT quantity FROM reagents WHERE id = ?', (reagent_id,))
        current_quantity = self.cursor.fetchone()[0]
        
        # 计算入库后的剩余库存
        remaining_quantity = current_quantity + quantity

        # 更新库存
        self.cursor.execute('''
            UPDATE reagents 
            SET quantity = quantity + ? 
            WHERE id = ?
        ''', (quantity, reagent_id))
        
        # 添加试剂名称到名称表（如果不存在）
        self.cursor.execute('''
            INSERT OR IGNORE INTO reagent_names (name) VALUES (?)
        ''', (name,))
        
        # 添加入库记录
        self.cursor.execute('''
            INSERT INTO inbound (reagent_id, reagent_name, reagent_batch, date, quantity, operator, gtin, remaining_quantity)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (reagent_id, name, batch, date, quantity, operator, gtin, remaining_quantity))
        self.conn.commit()
    
    def add_outbound(self, reagent_id, name, batch, date, quantity, operator):
        # 检查库存是否足够
        self.cursor.execute('SELECT quantity FROM reagents WHERE id = ?', (reagent_id,))
        current_quantity = self.cursor.fetchone()[0]
        
        if current_quantity < quantity:
            return False, "库存不足"
        
        # 计算出库后的剩余库存
        remaining_quantity = current_quantity - quantity
        # 更新库存
        self.cursor.execute('''
            UPDATE reagents 
            SET quantity = quantity - ? 
            WHERE id = ?
        ''', (quantity, reagent_id))
        
        # 添加出库记录（添加剩余库存字段）
        self.cursor.execute('''
            INSERT INTO outbound (reagent_id, reagent_name, reagent_batch, date, quantity, operator, gtin, remaining_quantity)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (reagent_id, name, batch, date, quantity, operator, None, remaining_quantity))
        self.conn.commit()
        return True, "出库成功"
    
    def add_new_reagent(self, name, batch, production_date, expiry_date, quantity, gtin=None, qrcode=None):
        try:
            # 确保试剂名称已存在
            self.cursor.execute('''
                INSERT OR IGNORE INTO reagent_names (name, gtin) VALUES (?, ?)
            ''', (name, gtin))

            # 添加试剂记录
            self.cursor.execute('''
                INSERT INTO reagents (name, batch, production_date, expiry_date, quantity, gtin, qrcode)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (name, batch, production_date, expiry_date, quantity, gtin, qrcode))
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            return False
        
    def update_reagent_name_and_gtin(self, old_name, new_name, old_gtin, new_gtin):
        """更新试剂名称和GTIN，并级联更新所有相关表"""
        try:
            # 开始事务
            self.conn.execute("BEGIN TRANSACTION")
            
            # 更新reagent_names表
            self.cursor.execute('''
                UPDATE reagent_names 
                SET name = ?, gtin = ?
                WHERE name = ? OR gtin = ?
            ''', (new_name, new_gtin, old_name, old_gtin))
            
            # 更新reagents表
            self.cursor.execute('''
                UPDATE reagents 
                SET name = ?, gtin = ?
                WHERE name = ? OR gtin = ?
            ''', (new_name, new_gtin, old_name, old_gtin))
            
            # 更新inbound表
            self.cursor.execute('''
                UPDATE inbound 
                SET reagent_name = ?, gtin = ?
                WHERE reagent_name = ? OR gtin = ?
            ''', (new_name, new_gtin, old_name, old_gtin))
            
            # 更新outbound表
            self.cursor.execute('''
                UPDATE outbound 
                SET reagent_name = ?, gtin = ?
                WHERE reagent_name = ? OR gtin = ?
            ''', (new_name, new_gtin, old_name, old_gtin))
            
            self.conn.commit()
            return True
        except sqlite3.Error as e:
            self.conn.rollback()
            print(f"更新失败: {str(e)}")
            return False

class DownloadThread(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(str, bool)
    error = pyqtSignal(str)

    def __init__(self, url, file_path, md5=None):
        super().__init__()
        self.url = url
        self.file_path = file_path
        self.md5 = md5

    def run(self):
        try:
            response = requests.get(self.url, stream=True, verify=False)
            total_size = int(response.headers.get('content-length', 0))
            downloaded = 0
            block_size = 1024
            
            with open(self.file_path, 'wb') as f:
                for data in response.iter_content(block_size):
                    f.write(data)
                    downloaded += len(data)
                    if total_size > 0:
                        progress = int(100 * downloaded / total_size)
                        self.progress.emit(progress)
            
            # 验证MD5
            if self.md5:
                file_md5 = self.calculate_md5(self.file_path)
                if file_md5 != self.md5:
                    self.error.emit("文件校验失败: MD5不匹配")
                    return
            
            self.finished.emit(self.file_path, True)
        except Exception as e:
            self.error.emit(f"下载失败: {str(e)}")

    def calculate_md5(self, file_path):
        hash_md5 = hashlib.md5()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_md5.update(chunk)
        return hash_md5.hexdigest()

class UpdateDialog(QDialog):
    def __init__(self,jsonData, parent=None):
        super().__init__(parent)
        self.setWindowTitle("软件更新")
        self.setFixedSize(400, 300)
        
        layout = QVBoxLayout()
        
        # 标题
        title_label = QLabel("软件更新")
        title_label.setFont(QFont("Arial", 12, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 状态信息
        self.status_label = QLabel(f"发现新版本 {jsonData['version']}")
        self.status_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(self.status_label)
        
        # 进度条
        self.progress_bar = QProgressBar()
        self.progress_bar.setRange(0, 100)
        self.progress_bar.setValue(0)
        layout.addWidget(self.progress_bar)
        
        # 更新日志
        self.changelog = QTextEdit()
        self.changelog.setReadOnly(True)
        self.changelog.setPlainText(jsonData.get('changelog', '无更新日志'))
        layout.addWidget(self.changelog)
        
        # 按钮布局
        button_layout = QHBoxLayout()
        self.update_btn = QPushButton("立即更新")
        self.update_btn.clicked.connect(self.start_update)
        button_layout.addWidget(self.update_btn)
        
        self.cancel_btn = QPushButton("取消")
        self.cancel_btn.clicked.connect(self.reject)
        button_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
        
        self.download_thread = None
        self.downloaded_file = None
        self.update_info = jsonData
    
    def start_update(self):
        if not self.update_info:
            return
            
        self.status_label.setText("正在下载更新...")
        self.update_btn.setEnabled(False)
        
        # 创建临时文件
        temp_dir = tempfile.gettempdir()
        self.downloaded_file = os.path.join(temp_dir, os.path.basename(self.update_info['url']))
        
        # 启动下载线程
        self.download_thread = DownloadThread(
            self.update_info['url'],
            self.downloaded_file,
            self.update_info.get('md5')
        )
        self.download_thread.progress.connect(self.progress_bar.setValue)
        self.download_thread.finished.connect(self.on_download_complete)
        self.download_thread.error.connect(self.on_download_error)
        self.download_thread.start()

    def on_download_complete(self, file_path, success):
        if success:
            self.status_label.setText("下载完成，准备安装...")
            self.install_update(file_path)
        else:
            self.status_label.setText("下载失败")

    def on_download_error(self, message):
        self.status_label.setText(message)
        self.update_btn.setEnabled(True)

    def install_update(self, file_path):
        try:
            save_path = os.path.join(DB_DIR, "saved_path.txt")
            print(save_path)
            # 写入路径到文件
            with open(save_path, 'w') as f:
                f.write(file_path)
                
            # 在Windows上使用静默安装
            subprocess.Popen([file_path, '/SILENT', '/RESTARTAPPLICATIONS'])
            
            self.accept()
            QApplication.quit()
        except Exception as e:
            # 如果安装失败，需要重新启动应用程序
            QMessageBox.warning(self.parent, "安装失败", f"安装过程中出错: {str(e)}")
            # 重启应用程序
            subprocess.Popen([sys.executable] + sys.argv)

class AutoCompleteComboBox(QComboBox):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setEditable(True)
        self.setInsertPolicy(QComboBox.NoInsert)
        self.completer().setCompletionMode(QCompleter.PopupCompletion)
        self.completer().setFilterMode(Qt.MatchContains)
        self.setMinimumContentsLength(0)
        
        # 存储完整列表
        self.full_list = []
        
        # 连接信号
        self.lineEdit().textEdited.connect(self.filter_items)
    
    def set_items(self, items, isNotClear=None):
        self.full_list = items
        
        # 更新ComboBox的项目列表（用于键盘导航）
        self.clear()
        self.addItems(items)

        # 关键修改：清空当前文本，避免默认选择第一项
        # if isNotClear:
            # self.setCurrentIndex(-1)  # 设置为-1表示没有选中项
            # self.setEditText('')      # 清空编辑框
    
    def filter_items(self, text):
        # 保存当前文本和光标位置
        current_text = self.lineEdit().text()
        cursor_pos = self.lineEdit().cursorPosition()
        
        # 过滤列表
        filtered = [item for item in self.full_list if text.lower() in item.lower()]
        
        # 更新下拉框
        self.clear()
        self.addItems(filtered)
        
        # 恢复文本和光标位置
        self.lineEdit().setText(current_text)
        self.lineEdit().setCursorPosition(cursor_pos)

class CombinedRecordPage(QWidget):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.date_range_enabled = False
        self.init_ui()
        self.load_data()
    
    def init_ui(self):
        layout = QVBoxLayout()

        barcode_layout = QHBoxLayout()
        # 添加扫码输入框
        self.barcode_input = QLineEdit()
        self.barcode_input.setPlaceholderText("请扫描试剂条码 (GS1标准)")
        self.barcode_input.setFixedWidth(500)
        self.barcode_input.textChanged.connect(self.process_barcode)

        barcode_clear_btn = QPushButton("清除")
        barcode_clear_btn.setFixedHeight(20)
        barcode_clear_btn.clicked.connect(self.clear_barcode)
        
        barcode_layout.addWidget(QLabel("条码扫描:"))
        barcode_layout.addWidget(self.barcode_input)
        barcode_layout.addWidget(barcode_clear_btn)
        barcode_layout.addStretch()
        
        # 创建搜索区域
        search_layout = QHBoxLayout()

        # 添加记录类型选择下拉框
        self.record_type_combo = QComboBox()
        self.record_type_combo.addItem("全部记录")
        self.record_type_combo.addItem("入库记录")
        self.record_type_combo.addItem("出库记录")
        self.record_type_combo.setFixedWidth(100)
        self.record_type_combo.currentIndexChanged.connect(self.load_data)
        
        self.name_search = QLineEdit()
        self.name_search.setPlaceholderText("试剂名称")
        self.name_search.setFixedWidth(200)
        self.name_search.textChanged.connect(self.load_data)
        
        self.batch_search = QLineEdit()
        self.batch_search.setPlaceholderText("批号")
        self.batch_search.setFixedWidth(150)
        self.batch_search.textChanged.connect(self.load_data)
        
        # 添加日期范围筛选
        date_layout = QHBoxLayout()
        date_layout.addWidget(QLabel("日期范围:"))
        
        # 开始日期选择器
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDate(QDate.currentDate().addMonths(-1))  # 默认一个月前
        self.start_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.start_date_edit.setDate(QDate())  # 初始设为空日期
        self.start_date_edit.dateChanged.connect(self.date_range_changed)
        date_layout.addWidget(self.start_date_edit)
        
        date_layout.addWidget(QLabel("至"))
        
        # 结束日期选择器
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDate(QDate.currentDate())  # 默认今天
        self.end_date_edit.setDisplayFormat("yyyy-MM-dd")
        self.end_date_edit.setDate(QDate())  # 初始设为空日期
        self.end_date_edit.dateChanged.connect(self.date_range_changed)
        date_layout.addWidget(self.end_date_edit)
        
        # 重置日期按钮
        reset_date_btn = QPushButton("重置日期")
        reset_date_btn.setFixedHeight(30)
        reset_date_btn.clicked.connect(self.reset_date_range)
        date_layout.addWidget(reset_date_btn)
        
        search_layout.addWidget(QLabel("记录类型:"))
        search_layout.addWidget(self.record_type_combo)
        search_layout.addWidget(QLabel("试剂名称:"))
        search_layout.addWidget(self.name_search)
        search_layout.addWidget(QLabel("批号:"))
        search_layout.addWidget(self.batch_search)
        search_layout.addLayout(date_layout)  # 添加日期布局

        search_layout.addStretch()
        
        layout.addLayout(barcode_layout)
        layout.addLayout(search_layout)

        # 创建表格视图
        self.table_view = QTableView()
        self.model = QStandardItemModel()
        
        # 设置合并后的表头
        self.model.setHorizontalHeaderLabels([
            "类型", "ID", "试剂名称", "批号", "日期", "数量", "操作员", "剩余库存"
        ])
            
        # 设置代理模型以支持排序
        self.proxy_model = QSortFilterProxyModel()
        self.proxy_model.setSourceModel(self.model)
        self.table_view.setModel(self.proxy_model)
        
        self.table_view.setSelectionBehavior(QTableView.SelectRows)
        self.table_view.setSortingEnabled(True)
        self.table_view.setAlternatingRowColors(True)
        self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.Interactive)
        self.table_view.verticalHeader().setDefaultSectionSize(30)

        # 设置列宽
        self.table_view.setColumnWidth(0, 60)   # 类型
        self.table_view.setColumnWidth(1, 50)   # ID
        self.table_view.setColumnWidth(2, 200)  # 试剂名称
        self.table_view.setColumnWidth(3, 120)  # 批号
        self.table_view.setColumnWidth(4, 220)  # 日期
        self.table_view.setColumnWidth(5, 90)   # 数量
        self.table_view.setColumnWidth(6, 80)   # 操作员
        self.table_view.setColumnWidth(7, 90)   # 剩余库存

        layout.addWidget(self.table_view)
        self.setLayout(layout)

    def clear_barcode(self):
        self.barcode_input.clear()
        self.name_search.clear()
        self.batch_search.clear()

    def process_barcode(self):
        """处理扫描的条码"""
        barcode = self.barcode_input.text().strip()
        if not barcode:
            return
            
        # 解析条码
        parsed = GS1Parser.parse_barcode(barcode)

        # 根据GTIN查找试剂名称
        if parsed.get('GTIN'):
            name = self.db.get_name_by_gtin(parsed.get('GTIN'))
            if name:
                self.name_search.setText(name)
        
        # 自动填充批号
        if parsed.get('Batch/Lot Number'):
            self.batch_search.setText(parsed.get('Batch/Lot Number'))

    def date_range_changed(self):
        """当日期范围改变时触发"""
        # 只有当至少一个日期被设置时才启用日期范围
        start_date_valid = self.start_date_edit.date().isValid()
        end_date_valid = self.end_date_edit.date().isValid()
        
        # 检查日期范围是否有效
        if start_date_valid and end_date_valid:
            if self.start_date_edit.date() > self.end_date_edit.date():
                QMessageBox.warning(self, "日期错误", "开始日期不能晚于结束日期")
                self.date_range_enabled = False
                return
        
        self.date_range_enabled = start_date_valid or end_date_valid
        self.load_data()
    
    def reset_date_range(self):
        """重置日期范围到默认值"""
        self.start_date_edit.setDate(QDate())  # 设为空日期
        self.end_date_edit.setDate(QDate())    # 设为空日期
        self.date_range_enabled = False
        self.load_data()
    
    def load_data(self):
        name_filter = self.name_search.text().strip()
        batch_filter = self.batch_search.text().strip()
        record_type = self.record_type_combo.currentText()
        
        # 获取日期范围（如果启用）
        start_date = None
        end_date = None
        if self.date_range_enabled:
            if self.start_date_edit.date().isValid():
                start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
            if self.end_date_edit.date().isValid():
                end_date = self.end_date_edit.date().toString("yyyy-MM-dd")
                end_date = end_date + " 23:59:59:999"

        # 根据选择的记录类型获取数据
        records = []
        if record_type == "全部记录" or record_type == "入库记录":
            inbound_records = self.db.get_inbound_records(
                name_filter, batch_filter, start_date, end_date
            )
            # 添加类型标记
            records.extend([("入库",) + rec for rec in inbound_records])
        
        if record_type == "全部记录" or record_type == "出库记录":
            outbound_records = self.db.get_outbound_records(
                name_filter, batch_filter, start_date, end_date
            )
            # 添加类型标记
            records.extend([("出库",) + rec for rec in outbound_records])
        
        # 按日期排序（从新到旧）
        records.sort(key=lambda x: x[4], reverse=True)
        
        # 更新表格数据
        self.model.setRowCount(0)
        for record in records:
            # record结构: (类型, id, 名称, 批号, 日期, 数量, 操作员, 剩余库存)
            type_item = QStandardItem(record[0])
            id_item = QStandardItem(str(record[1]))
            name_item = QStandardItem(str(record[2]))
            batch_item = QStandardItem(str(record[3]))
            date_item = QStandardItem(str(record[4]))
            quantity_item = QStandardItem(str(record[5]))
            operator_item = QStandardItem(str(record[6]))
            remaining_item = QStandardItem(str(record[7]))
            
            # 设置类型列的颜色
            if record[0] == "入库":
                type_item.setForeground(QBrush(QColor(0, 128, 0)))  # 绿色
            else:
                type_item.setForeground(QBrush(QColor(200, 0, 0)))  # 红色
                quantity_item.setForeground(QBrush(QColor(200, 0, 0)))  # 红色
            
            # 设置数量列对齐方式
            quantity_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            remaining_item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            
            # 设置日期列对齐方式
            date_item.setTextAlignment(Qt.AlignCenter)
            
            row_items = [
                type_item, id_item, name_item, batch_item, 
                date_item, quantity_item, operator_item, remaining_item
            ]
            
            self.model.appendRow(row_items)

class GS1Parser:
    """GS1条码解析器"""
    @staticmethod

    def parse_barcode(gs1_string):
        """解析GS1格式编码字符串
        
        参数:
        gs1_string (str): GS1格式编码字符串
        
        返回:
        dict: 包含解析后数据的字典
        """
        ai_list = [
            # 固定长度AI
            {"ai": "00", "name": "SSCC", "length": 18, "type": "numeric"},
            {"ai": "01", "name": "GTIN", "length": 14, "type": "numeric"},
            {"ai": "02", "name": "Content GTIN", "length": 14, "type": "numeric"},
            {"ai": "11", "name": "Production Date", "length": 6, "type": "numeric"},
            {"ai": "12", "name": "Due Date", "length": 6, "type": "numeric"},
            {"ai": "13", "name": "Packaging Date", "length": 6, "type": "numeric"},
            {"ai": "15", "name": "Best Before Date", "length": 6, "type": "numeric"},
            {"ai": "17", "name": "Expiration Date", "length": 6, "type": "numeric"},
            {"ai": "20", "name": "Product Variant", "length": 2, "type": "numeric"},
            {"ai": "30", "name": "Variable Count", "length": 8, "type": "numeric"},
            
            # 可变长度AI
            {"ai": "10", "name": "Batch/Lot Number", "length": None, "type": "alphanumeric"},
            {"ai": "21", "name": "Serial Number", "length": None, "type": "alphanumeric"},
            {"ai": "22", "name": "Consumer Product Variant", "length": None, "type": "alphanumeric"},
            {"ai": "240", "name": "Additional Item Identification", "length": None, "type": "alphanumeric"},
            {"ai": "241", "name": "Customer Part Number", "length": None, "type": "alphanumeric"},
            {"ai": "242", "name": "Made-to-Order Variation Number", "length": None, "type": "numeric"},
            {"ai": "250", "name": "Secondary Serial Number", "length": None, "type": "alphanumeric"},
            {"ai": "251", "name": "Reference to Source Entity", "length": None, "type": "alphanumeric"},
            {"ai": "253", "name": "Global Document Type Identifier", "length": 17, "type": "numeric"},
            {"ai": "254", "name": "GLN Extension Component", "length": None, "type": "alphanumeric"},
            {"ai": "255", "name": "Global Coupon Number", "length": 25, "type": "numeric"},
            
            # 带小数点的AI
            {"ai": "3100", "name": "Net Weight (kg)", "length": 6, "type": "numeric"},
            {"ai": "3101", "name": "Net Weight (kg)", "length": 6, "type": "numeric"},
            {"ai": "3102", "name": "Net Weight (kg)", "length": 6, "type": "numeric"},
            {"ai": "3103", "name": "Net Weight (kg)", "length": 6, "type": "numeric"},
            {"ai": "3104", "name": "Net Weight (kg)", "length": 6, "type": "numeric"},
            {"ai": "3105", "name": "Net Weight (kg)", "length": 6, "type": "numeric"},
            {"ai": "3920", "name": "Price", "length": None, "type": "numeric"},
            {"ai": "3921", "name": "Price", "length": None, "type": "numeric"},
            {"ai": "3922", "name": "Price", "length": None, "type": "numeric"},
            
            # 内部使用标识
            {"ai": "90", "name": "Internal", "length": None, "type": "alphanumeric"},
            {"ai": "91", "name": "Internal", "length": None, "type": "alphanumeric"},
            {"ai": "92", "name": "Internal", "length": None, "type": "alphanumeric"},
            {"ai": "93", "name": "Internal", "length": None, "type": "alphanumeric"},
            {"ai": "94", "name": "Internal", "length": None, "type": "alphanumeric"},
            {"ai": "95", "name": "Internal", "length": None, "type": "alphanumeric"},
            {"ai": "96", "name": "Internal", "length": None, "type": "alphanumeric"},
            {"ai": "97", "name": "Internal", "length": None, "type": "alphanumeric"},
            {"ai": "98", "name": "Internal", "length": None, "type": "alphanumeric"},
            {"ai": "99", "name": "Internal", "length": None, "type": "alphanumeric"},
        ]
                # 创建AI映射字典便于快速查找
        ai_map = {ai_info["ai"]: ai_info for ai_info in ai_list}
        gs1_string = gs1_string.replace("（", "(")
        gs1_string = gs1_string.replace("）", ")")
        # 正则表达式匹配括号内的AI和数据
        pattern = r'\((\d+)\)([^\(]*)'
        matches = re.findall(pattern, gs1_string)
        parsed_data = {}
        
        for ai_str, data in matches:
            # 查找匹配的AI定义
            ai_info = ai_map.get(ai_str)
            
            if not ai_info:
                # 未知AI，使用原始格式存储
                key = f"Unknown_AI({ai_str})"
                parsed_data[key] = data
                continue
                
            # 处理固定长度AI
            if ai_info["length"] is not None:
                # 检查数据长度是否符合预期
                if len(data) < ai_info["length"]:
                    # 数据不足，使用实际数据并记录错误
                    processed_data = data
                    key = f"ERROR_Short_{ai_info['name']}"
                elif len(data) > ai_info["length"]:
                    # 数据超长，截取所需部分
                    processed_data = data[:ai_info["length"]]
                    key = f"WARNING_Truncated_{ai_info['name']}"
                else:
                    processed_data = data
                    key = ai_info["name"]
                
                # 处理带小数点的AI (310X系列)
                if ai_str.startswith("310") and len(ai_str) == 4 and processed_data.isdigit():
                    decimal_pos = int(ai_str[3])
                    if decimal_pos > 0 and len(processed_data) > decimal_pos:
                        processed_data = processed_data[:-decimal_pos] + '.' + processed_data[-decimal_pos:]
            
            # 处理可变长度AI
            else:
                processed_data = data
                key = ai_info["name"]

            parsed_data[key] = processed_data
        
        return parsed_data

class ReagentNameManager(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle("试剂名称管理")
        self.setFixedSize(500, 400)
        
        layout = QVBoxLayout()

        # 标题
        title_label = QLabel("试剂名称管理")
        title_label.setFont(QFont("Arial", 12, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 添加搜索框
        search_layout = QHBoxLayout()
        
        self.search_edit = QLineEdit()
        self.search_edit.setPlaceholderText("搜索试剂名称...")
        self.search_edit.textChanged.connect(self.filter_reagent_names)
        
        search_layout.addWidget(QLabel("搜索:"))
        search_layout.addWidget(self.search_edit)
        
        layout.addLayout(search_layout)

        # 说明
        info_label = QLabel("在此管理系统中使用的试剂名称。添加新名称后，可在入库时使用。")
        info_label.setWordWrap(True)
        layout.addWidget(info_label)
        
        # 列表控件
        self.list_widget = QListWidget()
        self.list_widget.setAlternatingRowColors(True)
        layout.addWidget(self.list_widget)

        # 在表单布局中添加GTIN输入框
        form_layout = QFormLayout()
        
        self.barcode_input = QLineEdit()
        self.barcode_input.setPlaceholderText("请扫码")
        form_layout.addRow("二维码信息:", self.barcode_input)
        self.barcode_input.textChanged.connect(self.load_reagent_details)

        # 在表单布局中添加GTIN显示
        self.gtin_label = QLabel("")
        form_layout.addRow("GTIN:", self.gtin_label)

        self.new_name = QLineEdit()
        self.new_name.setPlaceholderText("输入新试剂名称")
        form_layout.addRow("新试剂名称:", self.new_name)
        
        # 按钮布局
        button_layout = QHBoxLayout()
        
        self.add_btn = QPushButton("添加新名称")
        self.add_btn.clicked.connect(self.add_reagent_name)
        form_layout.addWidget(self.add_btn)
        
        self.delete_btn = QPushButton("删除选中名称")
        self.delete_btn.clicked.connect(self.delete_reagent_name)
        
        # 添加编辑按钮
        self.edit_btn = QPushButton("编辑选中名称")
        self.edit_btn.clicked.connect(self.edit_reagent_name)
        
        button_layout.addWidget(self.delete_btn)
        button_layout.addWidget(self.edit_btn)
        
        layout.addLayout(button_layout)
        layout.addLayout(form_layout)
        
        # 关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignRight)
        
        self.setLayout(layout)
        
        # 加载试剂名称
        self.all_names = []  # 存储所有试剂名称
        self.load_reagent_names()

    def load_reagent_details(self):
        barcode = self.barcode_input.text().strip()

        parsed = GS1Parser.parse_barcode(barcode)
        # 显示GTIN
        if parsed.get('GTIN'):
            self.gtin_label.setText(parsed.get('GTIN') or "")
    
    def load_reagent_names(self):
        """加载所有试剂名称"""
        self.all_names = self.db.get_reagent_names()
        self.filter_reagent_names()
    
    def filter_reagent_names(self):
        """根据搜索文本过滤试剂名称"""
        search_text = self.search_edit.text().strip().lower()
        
        self.list_widget.clear()
        
        # 如果没有搜索文本，显示所有名称
        if not search_text:
            for name in self.all_names:
                item = QListWidgetItem(name)
                self.list_widget.addItem(item)
            return
        
        # 过滤匹配的试剂名称
        for name in self.all_names:
            if search_text in name.lower():
                item = QListWidgetItem(name)
                self.list_widget.addItem(item)
    
    def add_reagent_name(self):
        name = self.new_name.text().strip()
        if not name:
            QMessageBox.warning(self, "输入错误", "试剂名称不能为空")
            return
        
        gtin = self.gtin_label.text().strip() or None
        
        if self.db.add_reagent_name(name,gtin):
            # 更新本地列表
            self.all_names.append(name)
            self.all_names.sort()  # 保持排序
            self.filter_reagent_names()
            self.barcode_input.clear()
            self.gtin_label.clear()
            self.new_name.clear()

            QMessageBox.information(self, "成功", f"已添加试剂名称: {name}")
        else:
            QMessageBox.warning(self, "添加失败", f"试剂名称 '{name}' 已存在")
    
    def delete_reagent_name(self):
        selected_items = self.list_widget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "选择错误", "请先选择一个试剂名称")
            return
        
        name = selected_items[0].text()
        
        # 确认删除
        reply = QMessageBox.question(
            self, "确认删除", 
            f"确定要删除试剂名称 '{name}' 吗?\n注意: 如果该试剂已有库存或出入库记录，将无法删除。",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.Yes:
            success, message = self.db.delete_reagent_name(name)
            if success:
                # 更新本地列表
                if name in self.all_names:
                    self.all_names.remove(name)
                self.filter_reagent_names()
                QMessageBox.information(self, "成功", message)
            else:
                QMessageBox.warning(self, "删除失败", message)

    def edit_reagent_name(self):
        selected_items = self.list_widget.selectedItems()
        if not selected_items:
            QMessageBox.warning(self, "选择错误", "请先选择一个试剂名称")
            return
        
        name = selected_items[0].text()
        
        # 从数据库获取当前GTIN
        self.db.cursor.execute('SELECT gtin FROM reagent_names WHERE name = ?', (name,))
        result = self.db.cursor.fetchone()
        current_gtin = result[0] if result else ""
        
        # 创建编辑对话框
        dialog = ReagentEditDialog(name, current_gtin, self.db, self)
        if dialog.exec_() == QDialog.Accepted:
            # 刷新列表
            self.load_reagent_names()
            # 通知主窗口刷新数据
            if hasattr(self.parent(), 'load_data'):
                self.parent().load_data()
            if hasattr(self.parent(), 'combined_record_tab') and hasattr(self.parent().combined_record_tab, 'load_data'):
                self.parent().combined_record_tab.load_data()

# 添加新的编辑对话框类
class ReagentEditDialog(QDialog):
    def __init__(self, current_name, current_gtin, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.current_name = current_name
        self.current_gtin = current_gtin
        
        self.setWindowTitle("编辑试剂名称")
        self.setFixedSize(400, 200)
        
        layout = QVBoxLayout()
        
        # 标题
        title_label = QLabel("编辑试剂名称")
        title_label.setFont(QFont("Arial", 12, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 表单布局
        form_layout = QFormLayout()
        
        # 当前名称显示
        self.current_name_label = QLabel(current_name)
        form_layout.addRow("当前名称:", self.current_name_label)
        
        # 当前GTIN显示
        self.current_gtin_label = QLabel(current_gtin if current_gtin else "无")
        form_layout.addRow("当前GTIN:", self.current_gtin_label)
        
        # 扫码框
        self.barcode_input = QLineEdit()
        self.barcode_input.setPlaceholderText("不更新GTIN无需扫码")
        self.barcode_input.textChanged.connect(self.process_barcode)
        form_layout.addRow("扫码框:", self.barcode_input)
        
        # GTIN显示
        self.new_gtin_label = QLabel("")
        form_layout.addRow("新GTIN:", self.new_gtin_label)
        
        # 新名称输入
        self.new_name_edit = QLineEdit()
        self.new_name_edit.setPlaceholderText("输入新名称（留空则不修改）")
        form_layout.addRow("新名称:", self.new_name_edit)
        
        layout.addLayout(form_layout)
        
        # 按钮布局
        button_layout = QHBoxLayout()
        
        self.ok_btn = QPushButton("确定")
        self.ok_btn.clicked.connect(self.accept_edit)
        
        self.cancel_btn = QPushButton("关闭")
        self.cancel_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.ok_btn)
        button_layout.addWidget(self.cancel_btn)
        
        layout.addLayout(button_layout)
        self.setLayout(layout)
    
    def process_barcode(self):
        """处理扫描的条码"""
        barcode = self.barcode_input.text().strip()
        if not barcode:
            return
            
        # 解析条码
        parsed = GS1Parser.parse_barcode(barcode)
        
        # 显示GTIN
        if parsed.get('GTIN'):
            self.new_gtin_label.setText(parsed.get('GTIN') or "")
    
    def accept_edit(self):
        new_name = self.new_name_edit.text().strip()
        new_gtin = self.new_gtin_label.text().strip() or None
        
        # 如果没有提供新名称和新GTIN，则无需更新
        if not new_name and not new_gtin:
            QMessageBox.information(self, "提示", "未提供任何修改")
            return
        
        try:
            # 如果提供了新名称，检查是否与现有名称冲突（除了当前名称）
            if new_name and new_name != self.current_name:
                self.db.cursor.execute('SELECT COUNT(*) FROM reagent_names WHERE name = ?', (new_name,))
                if self.db.cursor.fetchone()[0] > 0:
                    QMessageBox.warning(self, "错误", f"试剂名称 '{new_name}' 已存在")
                    return
            
            # 如果提供了新GTIN，检查是否与现有GTIN冲突（除了当前GTIN）
            if new_gtin and new_gtin != self.current_gtin:
                self.db.cursor.execute('SELECT COUNT(*) FROM reagent_names WHERE gtin = ?', (new_gtin,))
                if self.db.cursor.fetchone()[0] > 0:
                    QMessageBox.warning(self, "错误", f"GTIN '{new_gtin}' 已被使用")
                    return
            
            # 使用新的方法更新所有相关表
            success = self.db.update_reagent_name_and_gtin(
                self.current_name, 
                new_name if new_name else self.current_name,  # 如果没有提供新名称，保持原名称
                self.current_gtin, 
                new_gtin if new_gtin else self.current_gtin,  # 如果没有提供新GTIN，保持原GTIN
            )
            
            if success:
                QMessageBox.information(self, "成功", "试剂信息已更新，所有相关记录已同步")
                self.accept()
            else:
                QMessageBox.critical(self, "错误", "更新失败，请检查数据库连接")
                
        except Exception as e:
            QMessageBox.critical(self, "错误", f"更新失败: {str(e)}")

class InboundDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setStyleSheet("""
            QDialog {
                background-color: #f5f7fa;
            }
            QLabel {
                font-weight: 500;
            }
        """)
        self.setWindowTitle("入库管理")
        self.setFixedSize(450, 350)

        layout = QVBoxLayout()
        
        form_layout = QFormLayout()

        # 添加扫码输入框
        self.barcode_input = QLineEdit()
        self.barcode_input.setPlaceholderText("请扫描试剂条码 (GS1标准)")
        self.barcode_input.textChanged.connect(self.process_barcode)
        form_layout.addRow("条码扫描:", self.barcode_input)

        # 在表单布局中添加GTIN显示
        self.gtin_label = QLabel("")
        form_layout.addRow("GTIN:", self.gtin_label)
        
        # 使用自定义的自动完成下拉框
        self.name_combo = AutoCompleteComboBox()
        self.name_combo.setEditable(True)
        self.name_combo.setPlaceholderText("输入试剂名称")
        self.name_combo.lineEdit().textChanged.connect(self.update_batches)
        self.name_combo.set_items(self.db.get_reagent_names())
        # self.name_combo.lineEdit().textEdited.connect(self.update_batches)
        form_layout.addRow("试剂名称:", self.name_combo)
        
        self.batch_combo = AutoCompleteComboBox()
        self.batch_combo.setPlaceholderText("输入批号")
        # self.batch_combo.lineEdit().textEdited.connect(self.on_batch_changed)
        form_layout.addRow("批号:", self.batch_combo)
        
        self.production_date_edit = QDateEdit()
        self.production_date_edit.setCalendarPopup(True)
        self.production_date_edit.setDate(QDate.currentDate().addMonths(-6))
        form_layout.addRow("生产日期:", self.production_date_edit)
        
        self.expiry_date_edit = QDateEdit()
        self.expiry_date_edit.setCalendarPopup(True)
        self.expiry_date_edit.setDate(QDate.currentDate().addMonths(12))
        form_layout.addRow("有效期至:", self.expiry_date_edit)
        
        self.quantity_edit = QLineEdit()
        self.quantity_edit.setPlaceholderText("输入入库数量")
        self.quantity_edit.setValidator(QIntValidator(1, 9999))
        form_layout.addRow("数量:", self.quantity_edit)
        
        self.operator_edit = QLineEdit()
        self.operator_edit.setPlaceholderText("输入操作员姓名")
        form_layout.addRow("操作员:", self.operator_edit)

        layout.addLayout(form_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept_inbound)
        button_box.rejected.connect(self.reject)
        
        layout.addWidget(button_box)
        self.setLayout(layout)
        
        # self.update_batches()

    
    def process_barcode(self):
        """处理扫描的条码"""
        barcode = self.barcode_input.text().strip()
        if not barcode:
            return
            
        # 解析条码
        parsed = GS1Parser.parse_barcode(barcode)
        # 显示GTIN
        # 根据GTIN查找试剂名称
        if parsed.get('GTIN'):
            self.gtin_label.setText(parsed.get('GTIN') or "")
            name = self.db.get_name_by_gtin(parsed.get('GTIN'))
            if name:
                self.name_combo.setCurrentText(name)
        
        # 自动填充批号
        if parsed.get('Batch/Lot Number'):
            self.batch_combo.lineEdit().setText(parsed.get('Batch/Lot Number'))
            
        # 自动填充生产日期
        if parsed.get('Production Date'):
            date_str = parsed.get('Production Date')
            year_part = date_str[:2]
            century = "20" if int(year_part) < 100 else "19"
            four_digit_year = century + year_part

            # 重构日期字符串
            new_date_str = four_digit_year + date_str[2:]

            date = QDate.fromString(new_date_str, "yyyyMMdd")
            if date.isValid():
                self.production_date_edit.setDate(date)
                
        # 自动填充有效期
        if parsed.get('Expiration Date'):
            date_str = parsed.get('Expiration Date')
            year_part = date_str[:2]
            century = "20" if int(year_part) < 100 else "19"
            four_digit_year = century + year_part

            # 重构日期字符串
            new_date_str = four_digit_year + date_str[2:]

            date = QDate.fromString(new_date_str, "yyyyMMdd")
            if date.isValid():
                self.expiry_date_edit.setDate(date)
                
        # 清空扫码框以便下次扫描
        # self.barcode_input.clear()
    
    def update_batches(self):
        current_name = self.name_combo.currentText()
        if current_name:
            batches = self.db.get_batches_by_name(current_name)
            self.batch_combo.set_items(batches)
        else:
            self.batch_combo.set_items([])
    
    # def on_batch_changed(self, text):
    #     # 如果批号被修改，尝试自动填充名称
    #     if text:
    #         name = self.db.get_name_by_batch(text)
    #         if name:
    #             self.name_combo.setCurrentText(name)
    
    def accept_inbound(self):
        name = self.name_combo.currentText().strip()
        gtin = self.gtin_label.text().strip() or None
        batch = self.batch_combo.currentText().strip()
        production_date = self.production_date_edit.date().toString("yyyy-MM-dd")
        expiry_date = self.expiry_date_edit.date().toString("yyyy-MM-dd")
        quantity = self.quantity_edit.text().strip()
        operator = self.operator_edit.text().strip()
        barcode = self.barcode_input.text().strip()

        if not name or not batch or not quantity:
            QMessageBox.warning(self, "输入错误", "请填写试剂名称、批号和数量")
            return
        
        try:
            quantity = int(quantity)
            if quantity <= 0:
                raise ValueError
        except ValueError:
            QMessageBox.warning(self, "输入错误", "请输入有效的数量（正整数）")
            return
        
        # 检查试剂是否已存在
        reagent_details = self.db.get_reagent_details(name, batch)
        
        if reagent_details:
            reagent_id = reagent_details[0]
        else:
            # 添加新试剂
            if not self.db.add_new_reagent(name, batch, production_date, expiry_date, 0, gtin, barcode):
                QMessageBox.warning(self, "操作失败", "添加新试剂失败，可能已存在相同名称和批号的试剂")
                return
            reagent_details = self.db.get_reagent_details(name, batch)
            reagent_id = reagent_details[0]
        
        # 执行入库
        self.db.add_inbound(
            reagent_id, 
            name, 
            batch, 
            QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss:zzz"), 
            quantity, 
            operator, 
            gtin
        )
        
        QMessageBox.information(self, "成功", f"试剂 {name} (批号: {batch}) 入库成功，数量: {quantity}")
        self.accept()

class OutboundDialog(QDialog):
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db

        self.setStyleSheet("""
            QDialog {
                background-color: #f5f7fa;
            }
            QLabel {
                font-weight: 500;
            }
        """)

        self.setWindowTitle("出库管理")
        self.setFixedSize(450, 350)
        
        layout = QVBoxLayout()
        
        form_layout = QFormLayout()

        # 添加扫码输入框
        self.barcode_input = QLineEdit()
        self.barcode_input.setPlaceholderText("请扫描试剂条码 (GS1标准)")
        self.barcode_input.textChanged.connect(self.process_barcode)
        form_layout.addRow("条码扫描:", self.barcode_input)
        
        # 在表单布局中添加GTIN显示
        self.gtin_label = QLabel("")
        form_layout.addRow("GTIN:", self.gtin_label)
        
        # 使用自定义的自动完成下拉框
        self.name_combo = AutoCompleteComboBox()
        self.name_combo.setEditable(True)
        self.name_combo.setPlaceholderText("输入试剂名称")
        self.name_combo.lineEdit().textChanged.connect(self.update_batches)
        self.name_combo.set_items(self.db.get_reagent_names())
        form_layout.addRow("试剂名称:", self.name_combo)
        
        self.batch_combo = AutoCompleteComboBox()
        self.batch_combo.setPlaceholderText("输入批号")
        # self.batch_combo.lineEdit().textEdited.connect(self.on_batch_changed)
        self.batch_combo.currentIndexChanged.connect(self.update_stock_info)
        form_layout.addRow("批号:", self.batch_combo)
        
        self.stock_label = QLabel("库存: 0")
        form_layout.addRow("当前库存:", self.stock_label)
        
        self.quantity_edit = QLineEdit()
        self.quantity_edit.setPlaceholderText("输入出库数量")
        self.quantity_edit.setValidator(QIntValidator(1, 9999))
        form_layout.addRow("数量:", self.quantity_edit)
        
        self.operator_edit = QLineEdit()
        self.operator_edit.setPlaceholderText("输入操作员姓名")
        form_layout.addRow("操作员:", self.operator_edit)

        layout.addLayout(form_layout)
        
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(self.accept_outbound)
        button_box.rejected.connect(self.reject)
        
        layout.addWidget(button_box)
        self.setLayout(layout)
        
        self.update_batches()
    
    def process_barcode(self):
        """处理扫描的条码"""
        barcode = self.barcode_input.text().strip()
        if not barcode:
            return
            
        # 解析条码
        parsed = GS1Parser.parse_barcode(barcode)
        
        # 根据GTIN查找试剂名称
        if parsed.get('GTIN'):
            self.gtin_label.setText(parsed.get('GTIN') or "")
            name = self.db.get_name_by_gtin(parsed.get('GTIN'))
            if name:
                self.name_combo.setCurrentText(name)
        
        # 自动填充批号
        if parsed.get('Batch/Lot Number'):
            self.batch_combo.lineEdit().setText(parsed.get('Batch/Lot Number'))
            self.update_stock_info()

    def update_batches(self):
        current_name = self.name_combo.currentText()
        if current_name:
            batches = self.db.get_batches_by_name(current_name)
            self.batch_combo.set_items(batches)
        else:
            self.batch_combo.set_items([])
    
    # def on_batch_changed(self, text):
    #     # 如果批号被修改，尝试自动填充名称
    #     if text:
    #         name = self.db.get_name_by_batch(text)
    #         if name:
    #             self.name_combo.setCurrentText(name)
    #             self.update_stock_info()
    
    def update_stock_info(self):
        name = self.name_combo.currentText().strip()
        batch = self.batch_combo.currentText().strip()
        
        if name and batch:
            reagent_details = self.db.get_reagent_details(name, batch)
            if reagent_details:
                self.stock_label.setText(f"库存: {reagent_details[3]}")
                return
        
        self.stock_label.setText("库存: 0")
    
    def accept_outbound(self):
        name = self.name_combo.currentText().strip()
        batch = self.batch_combo.currentText().strip()
        quantity = self.quantity_edit.text().strip()
        operator = self.operator_edit.text().strip()
        
        if not name or not batch or not quantity:
            QMessageBox.warning(self, "输入错误", "请填写试剂名称、批号和数量")
            return
        
        try:
            quantity = int(quantity)
            if quantity <= 0:
                raise ValueError
        except ValueError:
            QMessageBox.warning(self, "输入错误", "请输入有效的数量（正整数）")
            return
        
        # 获取试剂详情
        reagent_details = self.db.get_reagent_details(name, batch)
        if not reagent_details:
            QMessageBox.warning(self, "错误", "找不到指定的试剂")
            return
        
        reagent_id = reagent_details[0]
        
        # 执行出库
        success, message = self.db.add_outbound(
            reagent_id, 
            name, 
            batch, 
            QDateTime.currentDateTime().toString("yyyy-MM-dd HH:mm:ss:zzz"),
            quantity, 
            operator, 
        )
        
        if success:
            QMessageBox.information(self, "成功", f"试剂 {name} (批号: {batch}) 出库成功，数量: {quantity}")
            self.accept()
        else:
            QMessageBox.warning(self, "出库失败", message)

class MoreDialog(QDialog):
    """更多功能对话框，包含备份和导入数据库功能"""
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setWindowTitle("更多功能")
        self.setFixedSize(300, 350)
        
        layout = QVBoxLayout()
        layout.setSpacing(10)

        # 标题
        title_label = QLabel("数据库管理")
        title_label.setFont(QFont("Arial", 12, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 按钮布局
        button_layout = QVBoxLayout()
        button_layout.setSpacing(10)

        self.backup_btn = QPushButton("备份数据库")
        self.backup_btn.setFixedHeight(40)
        self.backup_btn.clicked.connect(self.backup_database)
        
        self.import_btn = QPushButton("导入数据库")
        self.import_btn.setFixedHeight(40)
        self.import_btn.clicked.connect(self.import_database)

        # 添加导出按钮
        self.export_btn = QPushButton("导出CSV")
        self.export_btn.setFixedHeight(40)
        self.export_btn.clicked.connect(self.export_records_to_xlsx)

        # 在按钮布局中添加更新按钮
        self.update_btn = QPushButton("更新")
        self.update_btn.setFixedHeight(40)
        self.update_btn.clicked.connect(self.check_for_updates)

        # 在按钮布局中添加关于按钮
        self.about_btn = QPushButton("关于")
        self.about_btn.setFixedHeight(40)
        self.about_btn.clicked.connect(self.show_about)
        
        button_layout.addWidget(self.backup_btn)
        button_layout.addWidget(self.import_btn)
        button_layout.addWidget(self.export_btn)
        button_layout.addWidget(self.update_btn)
        button_layout.addWidget(self.about_btn)

        layout.addLayout(button_layout)
        layout.addStretch()

        # 关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(self.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignRight)
        
        self.setLayout(layout)
    
    def backup_database(self):
        """备份数据库到用户选择的位置"""
        # 生成带时间戳的备份文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"ReagentWarehouseData_{timestamp}.db"
        
        # 让用户选择保存位置
        file_path, _ = QFileDialog.getSaveFileName(
            self, "备份数据库", default_filename, "数据库文件 (*.db)"
        )
        
        if not file_path:
            return  # 用户取消了操作
        
        try:
            # 复制数据库文件
            shutil.copyfile(DB_FILE, file_path)
            
            # 显示成功消息
            QMessageBox.information(
                self, 
                "备份成功", 
                f"数据库已成功备份到:\n{file_path}"
            )

            self.accept()

        except Exception as e:
            QMessageBox.critical(
                self, 
                "备份失败", 
                f"备份数据库时出错:\n{str(e)}"
            )
    
    def import_database(self):
        """导入新的数据库文件"""
        # 让用户选择数据库文件
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择数据库文件", "", "数据库文件 (*.db)"
        )
        
        if not file_path:
            return  # 用户取消了操作
        
        # 确认操作
        reply = QMessageBox.question(
            self, "确认导入", 
            "此操作将替换当前数据库并备份原有数据库。\n确定要继续吗？",
            QMessageBox.Yes | QMessageBox.No
        )
        
        if reply == QMessageBox.No:
            return
        
        try:
            # 备份当前数据库
            backup_dir = os.path.join(DB_DIR, "Backups")
            os.makedirs(backup_dir, exist_ok=True)

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = os.path.join(backup_dir, f"ReagentWarehouseData_backup_{timestamp}.db")
            shutil.copyfile(DB_FILE, backup_file)
            
            # 替换为新的数据库文件
            shutil.copyfile(file_path, DB_FILE)
            
            # 询问用户是否立即重启
            reply = QMessageBox.question(
                self, 
                "导入成功", 
                f"数据库已成功导入！\n原有数据库已备份到:\n{backup_file}\n\n是否立即重启应用？",
                QMessageBox.Yes | QMessageBox.No
            )

            if reply == QMessageBox.Yes:
                # 关闭所有窗口
                self.accept()
                # 重启应用
                QApplication.quit()
                os.execl(sys.executable, sys.executable, *sys.argv)

            else:
                # 显示成功消息
                QMessageBox.information(
                    self, 
                    "导入成功", 
                    f"数据库已成功导入！\n原有数据库已备份到:\n{backup_file}\n\n请稍后手动重启应用。"
                )
            
        except Exception as e:
            QMessageBox.critical(
                self, 
                "导入失败", 
                f"导入数据库时出错:\n{str(e)}"
            )

    def export_records_to_xlsx(self):
        """导出日期范围内的出入库记录到XLSX文件"""
        # 创建日期范围选择对话框
        dialog = QDialog(self)
        dialog.setWindowTitle("选择日期范围")
        dialog.setFixedSize(300, 150)
        
        layout = QVBoxLayout()
        form_layout = QFormLayout()
        
        # 开始日期选择器
        self.start_date_edit = QDateEdit()
        self.start_date_edit.setCalendarPopup(True)
        self.start_date_edit.setDate(QDate.currentDate().addMonths(-1))
        self.start_date_edit.setDisplayFormat("yyyy-MM-dd")
        form_layout.addRow("开始日期:", self.start_date_edit)
        
        # 结束日期选择器
        self.end_date_edit = QDateEdit()
        self.end_date_edit.setCalendarPopup(True)
        self.end_date_edit.setDate(QDate.currentDate())
        self.end_date_edit.setDisplayFormat("yyyy-MM-dd")
        form_layout.addRow("结束日期:", self.end_date_edit)
        
        layout.addLayout(form_layout)
        
        # 按钮框
        button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        button_box.accepted.connect(dialog.accept)
        button_box.rejected.connect(dialog.reject)
        
        layout.addWidget(button_box)
        dialog.setLayout(layout)
        
        if dialog.exec_() != QDialog.Accepted:
            return
        
        start_date = self.start_date_edit.date().toString("yyyy-MM-dd")
        end_date = self.end_date_edit.date().toString("yyyy-MM-dd")
        end_date = end_date + " 23:59:59:999"

        # 1. 获取所有在结束日期前有记录的试剂批号
        self.parent().db.cursor.execute('''
            SELECT DISTINCT reagent_name, reagent_batch
            FROM (
                SELECT reagent_name, reagent_batch
                FROM inbound
                WHERE date <= ?
                UNION
                SELECT reagent_name, reagent_batch
                FROM outbound
                WHERE date <= ?
            )
        ''', (end_date, end_date))
        distinct_reagents = self.parent().db.cursor.fetchall()

        # 2. 获取选定周期内的出入库记录
        inbound_records = self.parent().db.get_inbound_records("", "", start_date, end_date)
        outbound_records = self.parent().db.get_outbound_records("", "", start_date, end_date)
        
        # 合并记录并添加类型标记
        all_records = []
        for rec in inbound_records:
            # (id, reagent_name, reagent_batch, date, quantity, operator, remaining_quantity)
            all_records.append(('in', rec[1], rec[2], rec[3], rec[4], rec[5], rec[6]))
        
        for rec in outbound_records:
            all_records.append(('out', rec[1], rec[2], rec[3], rec[4], rec[5], rec[6]))
        
        # 按名称、批号和日期排序
        all_records.sort(key=lambda x: (x[1], x[2], x[3]))
        
        # 组织数据结构
        organized_data = {}

        # 为每个试剂批号准备数据结构
        for (name, batch) in distinct_reagents:
            key = (name, batch)
            organized_data[key] = {
                'initial_stock': 0,  # 初始库存（周期开始前最后一条记录的操作后库存）
                'records': [],        # 周期内记录
                'ending_stock': 0     # 周期末库存（周期结束前最后一条记录的操作后库存）
            }

        # 5. 获取每个试剂批号在周期开始前的最后一条记录（用于初始库存）
        for key, data in organized_data.items():
            name, batch = key
            
            # 获取周期开始前最后一条记录（入库或出库）
            self.parent().db.cursor.execute('''
                SELECT remaining_quantity, date, 'in' as type
                FROM inbound
                WHERE reagent_name = ? AND reagent_batch = ? AND date < ?
                UNION
                SELECT remaining_quantity, date, 'out' as type
                FROM outbound
                WHERE reagent_name = ? AND reagent_batch = ? AND date < ?
                ORDER BY date DESC
                LIMIT 1
            ''', (name, batch, start_date, name, batch, start_date))
            
            last_record_before_start = self.parent().db.cursor.fetchone()
            if last_record_before_start:
                data['initial_stock'] = last_record_before_start[0]

        # 6. 获取每个试剂批号在周期结束前的最后一条记录（用于周期末库存）
        for key, data in organized_data.items():
            name, batch = key
            
            # 获取周期结束前最后一条记录（入库或出库）
            self.parent().db.cursor.execute('''
                SELECT remaining_quantity, date, 'in' as type
                FROM inbound
                WHERE reagent_name = ? AND reagent_batch = ? AND date <= ?
                UNION
                SELECT remaining_quantity, date, 'out' as type
                FROM outbound
                WHERE reagent_name = ? AND reagent_batch = ? AND date <= ?
                ORDER BY date DESC
                LIMIT 1
            ''', (name, batch, end_date, name, batch, end_date))
            
            last_record_before_end = self.parent().db.cursor.fetchone()
            if last_record_before_end:
                data['ending_stock'] = last_record_before_end[0]

        # 7. 添加周期内记录
        for record in all_records:
            record_type, name, batch, date, quantity, operator, remaining = record
            key = (name, batch)
            
            if key in organized_data:
                organized_data[key]['records'].append({
                    'type': record_type,
                    'date': date,
                    'quantity': quantity,
                    'operator': operator,
                    'remaining': remaining
                })

        # 8. 过滤掉不需要的数据：初始库存为0，周期内没有记录，且周期末库存为0
        filtered_data = {}
        for key, data in organized_data.items():
            # 检查是否需要过滤
            if data['initial_stock'] == 0 and len(data['records']) == 0 and data['ending_stock'] == 0:
                continue  # 跳过这个试剂批号
            filtered_data[key] = data
        
        # 准备Excel数据
        excel_data = []
        
        # 添加标题行
        excel_data.append([
            "试剂名称", "试剂批号", "历史结余", "入库日期", "入库数量", 
            "出库日期", "出库数量", "操作后库存", "周期结余", "库存校验"
        ])
        
        # 添加数据行
        for (name, batch), data in filtered_data.items():
            # # 添加初始库存行
            # excel_data.append([
            #     name, batch, data['initial_stock'], "", "", "", "", "", 
            #     data['initial_stock'], data['ending_stock']
            # ])
            
            # 创建空记录，显示历史结余
            if len(data['records']) == 0:
                data['records'].append({
                    'type': 'in',
                    'date': '',
                    'quantity': '',
                    'operator': '',
                    'remaining': ''
                })

            # 添加记录行
            for record in data['records']:
                if record['type'] == 'in':
                    excel_data.append([
                        name, batch, data['initial_stock'], 
                        record['date'], record['quantity'], 
                        "", "", 
                        record['remaining'], 
                        data['ending_stock'],""
                    ])
                else:  # out
                    excel_data.append([
                        name, batch, data['initial_stock'], 
                        "", "", 
                        record['date'], record['quantity'], 
                        record['remaining'], 
                        data['ending_stock'],""
                    ])

        
        # 提示用户选择保存位置
        dateStr = QDate.currentDate().toString("yyyy-MM-dd")

        name_start_date = self.start_date_edit.date().toString("yy-MM-dd")
        name_end_date = self.end_date_edit.date().toString("yy-MM-dd")
        xlsx_name = f"出入库记录_{name_start_date}_至_{name_end_date}.xlsx"
        
        file_path, _ = QFileDialog.getSaveFileName(
            self, "导出出入库记录", xlsx_name, "Excel文件 (*.xlsx)"
        )
        
        if not file_path:
            return
        
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "出入库记录"
            
            # 添加数据
            for row in excel_data:
                ws.append(row)
            
            # 设置样式 - 加粗标题行
            bold_font = Font(bold=True)
            for cell in ws[1]:
                cell.font = bold_font
            
            # 设置列宽
            column_widths = [20, 20, 10, 15, 10, 15, 10, 10, 10, 10]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[get_column_letter(i)].width = width
            
            # 设置居中对齐
            center_aligned = Alignment(horizontal='center', vertical='center')
            for row in ws.iter_rows(min_row=1, max_row=len(excel_data)):
                for cell in row:
                    cell.alignment = center_aligned
            
            # 合并单元格
            current_name = None
            name_start_row = 2
            current_batch = None
            batch_start_row = 2
            
            for row_idx in range(2, len(excel_data) + 1):
                name = ws.cell(row=row_idx, column=1).value
                batch = ws.cell(row=row_idx, column=2).value

                if not current_name:
                    current_name = name
                if not current_batch:
                    current_batch = batch

                # 合并批号单元格
                if name and batch:
                    if (batch != current_batch) or (name != current_name) and row_idx > batch_start_row:
                        ws.merge_cells(f'B{batch_start_row}:B{row_idx-1}')
                        ws.merge_cells(f'C{batch_start_row}:C{row_idx-1}')
                        ws.merge_cells(f'I{batch_start_row}:I{row_idx-1}')
                        ws.merge_cells(f'J{batch_start_row}:J{row_idx-1}')
                        current_batch = batch
                        batch_start_row = row_idx

                # 合并名称单元格
                if name:
                    if name != current_name and row_idx > name_start_row:
                        ws.merge_cells(f'A{name_start_row}:A{row_idx-1}')
                        current_name = name
                        name_start_row = row_idx
                                
                # 最后一行处理
                if row_idx == len(excel_data):
                    if name == current_name:
                        ws.merge_cells(f'A{name_start_row}:A{row_idx}')
                    if batch == current_batch:
                        ws.merge_cells(f'B{batch_start_row}:B{row_idx}')
                        ws.merge_cells(f'C{batch_start_row}:C{row_idx}')
                        ws.merge_cells(f'I{batch_start_row}:I{row_idx}')
                        ws.merge_cells(f'J{batch_start_row}:J{row_idx}')
            
            # 保存文件
            wb.save(file_path)
            QMessageBox.information(self, "导出成功", f"出入库记录已成功导出到: {file_path}")
            self.accept()
        except PermissionError as e:
            QMessageBox.critical(
                self, 
                "导出失败", 
                f"无法保存文件，请检查：\n"
                f"1. 目标位置是否有写入权限\n"
                f"2. 文件是否被其他程序打开\n"
                f"3. 是否选择了正确的保存位置\n\n"
                f"错误详情: {str(e)}"
            )
        except Exception as e:
            QMessageBox.critical(self, "导出失败", f"导出数据时出错: {str(e)}")

    def check_for_updates(self):
        try:
            config = configparser.ConfigParser()
            config_path = os.path.join(application_path, 'config.ini')

            # 检查文件是否存在
            if not os.path.exists(config_path):
                QMessageBox.information(None, "软件更新", "配置缺失，无法链接服务器")
                return None

            config.read(config_path)
            update_url = config.get('Update', 'Server', fallback='')

            if not update_url:
                QMessageBox.information(None, "软件更新", "未配置更新服务器")
                return None

            # 添加版本号作为查询参数
            current_version = config.get('Application', 'Version', fallback='1.0.0')
            response = requests.get(f"{update_url}/update.json?version={current_version}", timeout=5, verify=False)
            response.raise_for_status()

            update_info = response.json()
            if update_info['version'] > current_version:
                dialog = UpdateDialog(update_info, self)
                dialog.exec_()
            else:
                QMessageBox.information(self, "软件更新", "当前已是最新版本")
                return None
        except Exception as e:
            QMessageBox.information(self, "软件更新", "当前已是最新版本")
            return None

    def show_about(self):
        """显示关于对话框"""
        # 获取应用版本信息
        config = configparser.ConfigParser()
        config_path = os.path.join(application_path, 'config.ini')
        config.read(config_path)
        version = config.get('Application', 'Version', fallback='1.0.0')
        
        # 创建关于对话框
        about_dialog = QDialog(self)
        about_dialog.setWindowTitle("关于试剂库存管理系统")
        about_dialog.setFixedSize(500, 400)
        
        layout = QVBoxLayout()
        
        # 标题
        title_label = QLabel("试剂库存管理系统")
        title_label.setFont(QFont("Arial", 14, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 版本信息
        version_label = QLabel(f"版本: {version}")
        version_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(version_label)
        
        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        layout.addWidget(line)
        
        # 说明文本
        about_text = QTextEdit()
        about_text.setReadOnly(True)
        about_text.setHtml("""
        <h1>开源声明</h1>
        <h2>软件许可证</h2>
        <p>本软件使用 <strong>GPL v3 许可证</strong>。</p>
        <p>这意味着您拥有以下权利：</p>
        <ul>
            <li>自由运行本软件，无论出于何种目的</li>
            <li>自由学习并修改软件源代码</li>
            <li>自由分发软件副本</li>
            <li>自由分发您修改后的软件版本</li>
        </ul>
        
        <h2>依赖框架</h2>
        <p>本软件基于以下开源框架构建：</p>
        <ul>
            <li>
                <p><strong>PyQt5</strong> - 采用 GPL v3 许可证</p>
                <p>官网：<a href="https://www.riverbankcomputing.com/software/pyqt/">https://www.riverbankcomputing.com/software/pyqt/</a></p>
            </li>
        </ul>
        
        <h2>源代码</h2>
        <p>根据 GPL v3 要求，本软件源代码已公开：</p>
        <ul>
            <li>
                <p>GitHub 仓库：</p>
                <p><a href="https://github.com/WeijieG/ReagentManagementSystem.git">https://github.com/WeijieG/ReagentManagementSystem.git</a></p>
            </li>
        </ul>
        <p>您可以在该仓库中：</p>
        <ul>
            <li>获取完整的项目源代码</li>
            <li>查看项目文档和构建说明</li>
            <li>提交问题报告或建议</li>
            <li>创建分支并贡献代码</li>
        </ul>
        
        <h2>GPL v3 要求</h2>
        <p>根据许可证要求，当您分发本软件或修改版本时，必须：</p>
        <ul>
            <li>提供完整的源代码</li>
            <li>使用相同的许可证条款（GPL v3）</li>
            <li>明确版权声明和免责条款</li>
            <li>向接收者提供获取源代码的方式</li>
        </ul>
        
        <h2>用户权利</h2>
        <p>根据 GPL v3 许可证，您有权：</p>
        <ul>
            <li>自由获取、修改和再分发源代码</li>
            <li>将软件用于商业目的</li>
            <li>创建并分发软件的修改版本</li>
            <li>在遵守许可证条款的前提下，将本软件整合到其他项目中</li>
        </ul>
    
        <h1>项目详情</h1>

        <h2>系统说明</h2>
        <p>试剂库存管理系统是一个专业的实验室试剂管理解决方案，旨在帮助实验室管理人员高效、准确地管理试剂库存。</p>
        
        <h2>主要功能</h2>
        <ul>
            <li>试剂入库管理</li>
            <li>试剂出库管理</li>
            <li>库存实时监控</li>
            <li>周期数据导出</li>
            <li>有效期预警</li>
            <li>出入库记录查询</li>
            <li>条码扫描支持</li>
            <li>数据备份与恢复</li>
            <li>联网更新</li>
        </ul>

        <h2>功能说明书</h2>
        <ul>
            <li>试剂名称管理</li>
            <p>             对于未匹配过的试剂名，需要手动输入试剂名称，可以在试剂名称管理输入，也可以直接入库，会自动记录。对于GTIN编码匹配过的试剂名称，扫码即可自动填充</p>
            <li>试剂入库管理</li>
            <p>             对于标准GS1二维码，可以自动填充二维码内包含的生产日期，有效期，批号，非标准二维码需要手填输入，操作员一栏选填，填了便于确定谁操作的</p>
            <li>试剂出库管理</li>
            <p>             对于标准GS1二维码，可以自动填充二维码内包含的批号，非标准二维码需要手填输入，操作员一栏选填，填了便于确定谁操作的</p>
            <li>库存实时监控</li>
            <li>周期数据导出</li>
            <li>有效期预警</li>
            <li>出入库记录查询</li>
            <li>条码扫描支持</li>
            <li>数据备份与恢复</li>
            <p>             更多功能-数据库备份/数据库导入 支持SQLite轻量级数据库</p>
            <li>联网更新</li>
        </ul>
        
        <h1>注</h1>
        <ul>
            <li>本系统对GS1解析，依赖输入文本自动解析分隔符并添加的括号，不排除换把扫码枪就用不了的可能。</li>
            <li>理论上符合 GS1-128、GS1 DataBar Omnidirectional、GS1 DataBar Limited，GS1 DataBar Expanded标准的扫码枪，应该都能用，前提是扫码枪设置正确。</li>
            <li>二维码需要符合GS1标准，才能获取批号、日期等数据，不排除某些非标准的二维码扫码无效。</li>
            <li>如有需求变更或者功能需要修改，请找最初的负责人，名字留着配置里了。</li>
            <li>系统有本地完整性校验功能，如出现文件丢失，请立即停止使用，以防应用异常导致数据丢失（ps：能看到这里证明目前应用正常）</li>
        </ul>
        """)
        layout.addWidget(about_text)
        
        # 关闭按钮
        close_btn = QPushButton("关闭")
        close_btn.clicked.connect(about_dialog.accept)
        layout.addWidget(close_btn, alignment=Qt.AlignRight)
        
        about_dialog.setLayout(layout)
        about_dialog.exec_()

# 添加免责声明对话框类
class DisclaimerDialog(QDialog):
    def __init__(self, parent=None):
        super().__init__(parent)

        self.setWindowTitle("用前须知")
        self.setFixedSize(600, 500)
        
        layout = QVBoxLayout()
        
        # 标题
        title_label = QLabel("用前须知")
        title_label.setFont(QFont("Arial", 14, QFont.Bold))
        title_label.setAlignment(Qt.AlignCenter)
        layout.addWidget(title_label)
        
        # 分隔线
        line = QFrame()
        line.setFrameShape(QFrame.HLine)
        line.setFrameShadow(QFrame.Sunken)
        layout.addWidget(line)
        
        # 声明文本
        disclaimer_text = QTextEdit()
        disclaimer_text.setReadOnly(True)
        disclaimer_text.setFont(QFont("Arial", 10))
        disclaimer_text.setPlainText(DISCLAIMER_TEXT)
        layout.addWidget(disclaimer_text)
        
        # 确认复选框
        self.accept_checkbox = QCheckBox("我已阅读、理解并同意接受上述条款")
        layout.addWidget(self.accept_checkbox)
        
        # 按钮布局
        button_layout = QHBoxLayout()
        self.accept_btn = QPushButton("接受")
        self.accept_btn.setEnabled(False)
        self.accept_btn.clicked.connect(self.accept)
        
        self.reject_btn = QPushButton("拒绝")
        self.reject_btn.clicked.connect(self.reject)
        
        button_layout.addWidget(self.accept_btn)
        button_layout.addWidget(self.reject_btn)

        layout.addLayout(button_layout)
        
        self.setLayout(layout)
        
        # 连接复选框状态改变信号
        self.accept_checkbox.stateChanged.connect(self.update_accept_button)
    
    def update_accept_button(self, state):
        """根据复选框状态更新接受按钮状态"""
        self.accept_btn.setEnabled(state == Qt.Checked)

class ReagentManagementSystem(QMainWindow):
    def __init__(self):
        super().__init__()
        self.db = ReagentDatabase()
        self.current_page = 1
        self.page_size = 15

        # 检查免责声明
        if not self.check_disclaimer():
            sys.exit(0)
        
        # 检查文件完整性
        if not self.check_file_integrity():
            sys.exit(0)

        self.init_ui()
        self.load_data()
        self.setWindowIcon(QIcon(ICON_PATH))
        self.setWindowTitle("试剂库存管理系统")

        self.cleanup_old_installer()

    def check_file_integrity(self):
        """检查关键文件完整性"""
        config = configparser.ConfigParser()
        config_path = os.path.join(application_path, 'config.ini')
        
        # 如果配置文件不存在，无法进行校验
        if not os.path.exists(config_path):
            QMessageBox.critical(
                None, 
                "文件完整性错误", 
                "配置文件不存在，无法进行完整性校验。"
            )
            return False
        
        config.read(config_path)
        
        # 尝试获取存储的哈希值
        stored_hashe_86 = ""
        stored_hashe_64 = ""
        if config.has_option('Security', 'MainFileHashe_x86'):
            stored_hashe_86 = json.loads(config.get('Security', 'MainFileHashe_x86'))

        if config.has_option('Security', 'MainFileHashe_x64'):
            stored_hashe_64 = json.loads(config.get('Security', 'MainFileHashe_x64'))
        
        # 计算当前文件的哈希值
        current_hashes = self.calculate_file_hashes()
        # print("current_hashes",current_hashes)
        if stored_hashe_86 != current_hashes and stored_hashe_64 != current_hashes and getattr(sys, 'frozen', False):
            QMessageBox.critical(
                None, 
                "应用不完整！", 
                "应用完整性异常，出于安全考虑，程序将退出。"
            )
            return False
        
        return True
        

    def calculate_file_hashes(self):
        """计算关键文件的哈希值"""
        
        combined_hash = hashlib.sha256()

        file_name = 'ReagentManagementSystem.exe'
        file_path = os.path.join(application_path, file_name)
        
        if not os.path.exists(file_path):
            return '-1'
        
        try:
            with open(file_path, 'rb') as f:
                chunk = f.read(8192)
                while chunk:
                    combined_hash.update(chunk)
                    chunk = f.read(8192)
        except Exception as e:
            QMessageBox.warning(
                None,
                "文件读取错误",
                f"无法读取文件 {file_name}: {str(e)}"
            )
            return -1
        
        return combined_hash.hexdigest()

    def check_disclaimer(self):
        """检查用户是否已接受免责声明"""
        userData = configparser.ConfigParser()
        userData_path = os.path.join(DB_DIR, 'userData.ini')
        
        # 如果配置文件不存在则创建
        if not os.path.exists(userData_path):
            userData['General'] = {'DisclaimerAccepted': 'False'}
            with open(userData_path, 'w') as configfile:
                userData.write(configfile)
        
        userData.read(userData_path)
        
        # 检查是否已接受免责声明
        if userData.get('General', 'DisclaimerAccepted', fallback='False') == 'True':
            return True
            
        # 显示免责声明对话框
        dialog = DisclaimerDialog(self)
        result = dialog.exec_()
        
        if result == QDialog.Accepted:
            # 更新配置文件
            userData.set('General', 'DisclaimerAccepted', 'True')
            with open(userData_path, 'w') as configfile:
                userData.write(configfile)
            return True
        return False

    def init_ui(self):
        self.setWindowTitle("ReagentManagementSystem")
        self.setGeometry(100, 100, 1000, 700)

        main_widget = QWidget()
        main_layout = QVBoxLayout()
        
        # 创建顶部按钮组
        button_layout = QHBoxLayout()
        button_layout.setSpacing(10)
        button_layout.setContentsMargins(5, 5, 5, 5)

        self.inbound_btn = QPushButton("入库管理")
        self.inbound_btn.setFixedHeight(50)
        self.inbound_btn.setFixedWidth(120)
        self.inbound_btn.setStyleSheet("font-size: 15px;")
        self.inbound_btn.clicked.connect(self.show_inbound_dialog)
        self.inbound_btn.setStyleSheet(f"""
            QPushButton {{
                border: none;
                background-image: url({get_image_path('normal.png')});
                background-repeat: no-repeat;
                background-position: center;
            }}
            QPushButton:hover {{
                background-image: url({get_image_path('hover.png')});
            }}
            QPushButton:pressed {{
                background-image: url({get_image_path('pressed.png')});
            }}
        """)

        self.outbound_btn = QPushButton("出库管理")
        self.outbound_btn.setFixedHeight(50)
        self.outbound_btn.setFixedWidth(120)
        self.outbound_btn.setStyleSheet("font-size: 15px;")
        self.outbound_btn.clicked.connect(self.show_outbound_dialog)
        self.outbound_btn.setStyleSheet(f"""
            QPushButton {{
                border: none;
                background-image: url({get_image_path('normal.png')});
                background-repeat: no-repeat;
                background-position: center;
            }}
            QPushButton:hover {{
                background-image: url({get_image_path('hover.png')});
            }}
            QPushButton:pressed {{
                background-image: url({get_image_path('pressed.png')});
            }}
        """)
        
        self.name_manager_btn = QPushButton("试剂名称管理")
        self.name_manager_btn.setFixedHeight(50)
        self.name_manager_btn.setFixedWidth(120)
        self.name_manager_btn.setStyleSheet("font-size: 15px;")
        self.name_manager_btn.clicked.connect(self.show_name_manager)
        self.name_manager_btn.setStyleSheet(f"""
            QPushButton {{
                border: none;
                background-image: url({get_image_path('normal.png')});
                background-repeat: no-repeat;
                background-position: center;
            }}
            QPushButton:hover {{
                background-image: url({get_image_path('hover.png')});
            }}
            QPushButton:pressed {{
                background-image: url({get_image_path('pressed.png')});
            }}
        """)

        # 新增"更多"按钮
        self.more_btn = QPushButton("更多功能")
        self.more_btn.setFixedHeight(50)
        self.more_btn.setFixedWidth(120)
        self.more_btn.setStyleSheet("font-size: 15px;")
        self.more_btn.clicked.connect(self.show_more_dialog)
        self.more_btn.setStyleSheet(f"""
            QPushButton {{
                border: none;
                background-image: url({get_image_path('normal.png')});
                background-repeat: no-repeat;
                background-position: center;
            }}
            QPushButton:hover {{
                background-image: url({get_image_path('hover.png')});
            }}
            QPushButton:pressed {{
                background-image: url({get_image_path('pressed.png')});
            }}
        """)
        
        button_layout.addWidget(self.inbound_btn)
        button_layout.addWidget(self.outbound_btn)
        button_layout.addWidget(self.name_manager_btn)
        button_layout.addStretch()
        button_layout.addWidget(self.more_btn)  # 添加"更多"按钮

        main_layout.addLayout(button_layout)
        
        # 创建选项卡
        self.tab_widget = QTabWidget()
        
        # 库存管理标签页
        self.inventory_tab = QWidget()
        inventory_layout = QVBoxLayout()

        barcode_layout = QHBoxLayout()
        # 添加扫码输入框
        self.barcode_input = QLineEdit()
        self.barcode_input.setPlaceholderText("请扫描试剂条码 (GS1标准)")
        self.barcode_input.setFixedWidth(500)
        self.barcode_input.textChanged.connect(self.process_barcode)

        barcode_clear_btn = QPushButton("清除")
        barcode_clear_btn.setFixedHeight(20)
        barcode_clear_btn.clicked.connect(self.clear_barcode)
        
        # barcode_layout.addRow("条码扫描:", self.barcode_input)
        barcode_layout.addWidget(QLabel("条码扫描:"))
        barcode_layout.addWidget(self.barcode_input)
        barcode_layout.addWidget(barcode_clear_btn)
        barcode_layout.addStretch()

        
        # 创建搜索区域
        search_layout = QHBoxLayout()
        
        self.name_search = QLineEdit()
        self.name_search.setPlaceholderText("试剂名称")
        self.name_search.setFixedWidth(200)
        
        self.batch_search = QLineEdit()
        self.batch_search.setPlaceholderText("批号")
        self.batch_search.setFixedWidth(150)
        
        search_btn = QPushButton("搜索")
        search_btn.clicked.connect(self.load_data)
        
        search_layout.addWidget(QLabel("试剂名称:"))
        search_layout.addWidget(self.name_search)
        search_layout.addWidget(QLabel("批号:"))
        search_layout.addWidget(self.batch_search)
        search_layout.addWidget(search_btn)
        search_layout.addStretch()

        inventory_layout.addLayout(barcode_layout)
        inventory_layout.addLayout(search_layout)
        
        # 创建表格视图
        self.table_view = QTableView()
        # 设置高度策略，允许表格根据需要扩展并显示滚动条
        self.table_view.setSizePolicy(
            QSizePolicy.Expanding, 
            QSizePolicy.Expanding
        )
        self.model = QStandardItemModel()
        self.model.setHorizontalHeaderLabels([
            "ID", "试剂名称", "批号", "生产日期", "有效期至", "库存数量", "剩余有效天数", "GTIN"  # 新增GTIN列
        ])
        self.table_view.setModel(self.model)
        self.table_view.setSelectionBehavior(QTableView.SelectRows)
        # 不占满，运行调整
        # self.table_view.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table_view.setAlternatingRowColors(True)

        # 设置列宽（调整列宽以适应新列）
        self.table_view.setColumnWidth(0, 50)   # ID
        self.table_view.setColumnWidth(1, 200)  # 试剂名称
        self.table_view.setColumnWidth(2, 150)  # 批号
        self.table_view.setColumnWidth(3, 100)  # 生产日期
        self.table_view.setColumnWidth(4, 100)  # 有效期至
        self.table_view.setColumnWidth(5, 80)   # 库存数量
        self.table_view.setColumnWidth(6, 100)  # 剩余有效天数
        self.table_view.setColumnWidth(7, 150)  # GTIN列

        self.table_view.verticalHeader().setDefaultSectionSize(30)
        
        inventory_layout.addWidget(self.table_view)
        
        # 创建分页控件
        page_layout = QHBoxLayout()
        
        self.prev_btn = QPushButton("上一页")
        self.prev_btn.clicked.connect(self.prev_page)
        
        self.next_btn = QPushButton("下一页")
        self.next_btn.clicked.connect(self.next_page)
        
        self.page_label = QLabel("第 1 页")

        # 添加每页显示条数选择框
        page_layout.addWidget(QLabel("每页显示:"))

        self.page_size_combo = QComboBox()
        self.page_size_combo.addItems(["30", "50", "100"])
        self.page_size_combo.setCurrentText("10")
        self.page_size_combo.setFixedWidth(80)
        self.page_size_combo.currentTextChanged.connect(self.change_page_size)
        page_layout.addWidget(self.page_size_combo)
        
        page_layout.addStretch()
        page_layout.addWidget(self.prev_btn)
        page_layout.addWidget(self.page_label)
        page_layout.addWidget(self.next_btn)
        page_layout.addStretch()
        
        inventory_layout.addLayout(page_layout)
        
        self.inventory_tab.setLayout(inventory_layout)
        # 合并的出入库记录标签页
        self.combined_record_tab = CombinedRecordPage(self.db)
        # 添加选项卡
        self.tab_widget.addTab(self.inventory_tab, "库存管理")
        self.tab_widget.addTab(self.combined_record_tab, "出入库记录")
        
        main_layout.addWidget(self.tab_widget)

        config = configparser.ConfigParser()
        config_path = os.path.join(application_path, 'config.ini')
        config.read(config_path)
        version = config.get('Application', 'Version', fallback='1.0.0')
        
        # 状态栏
        self.status_bar = QStatusBar()
        self.setStatusBar(self.status_bar)
        self.status_bar.showMessage(f"数据库文件: {DB_FILE} | 就绪")
        self.status_bar.showMessage(f"就绪 | 数据库: {os.path.basename(DB_FILE)} | 版本: {version}")

        main_widget.setLayout(main_layout)
        self.setCentralWidget(main_widget)

    def clear_barcode(self):
        self.barcode_input.clear()
        self.name_search.clear()
        self.batch_search.clear()

    def process_barcode(self):
        """处理扫描的条码"""
        barcode = self.barcode_input.text().strip()
        if not barcode:
            return
            
        # 解析条码
        print(barcode)
        parsed = GS1Parser.parse_barcode(barcode)

        # 根据GTIN查找试剂名称
        if parsed.get('GTIN'):
            name = self.db.get_name_by_gtin(parsed.get('GTIN'))
            if name:
                self.name_search.setText(name)
        
        # 自动填充批号
        if parsed.get('Batch/Lot Number'):
            self.batch_search.setText(parsed.get('Batch/Lot Number'))

    def show_more_dialog(self):
        """显示更多功能对话框"""
        dialog = MoreDialog(self)
        dialog.exec_()

    def change_page_size(self):
        """更改每页显示条数"""
        try:
            # 获取新的每页条数
            new_page_size = int(self.page_size_combo.currentText())
            self.page_size = new_page_size
            
            # 重置到第一页
            self.current_page = 1
            
            # 重新加载数据
            self.load_data()
        except ValueError:
            # 处理无效值
            QMessageBox.warning(self, "错误", "无效的每页显示条数")
            self.page_size_combo.setCurrentText(str(self.page_size))

    def load_data(self):
        name_filter = self.name_search.text().strip()
        batch_filter = self.batch_search.text().strip()
        
        # 获取总记录数
        query = "SELECT COUNT(*) FROM reagents WHERE 1=1 AND quantity > 0"
        params = []
        
        if name_filter:
            query += " AND name LIKE ?"
            params.append(f"%{name_filter}%")
        
        if batch_filter:
            query += " AND batch LIKE ?"
            params.append(f"%{batch_filter}%")
        
        self.db.cursor.execute(query, params)
        total_records = self.db.cursor.fetchone()[0]
        
        # 计算总页数
        total_pages = max(1, (total_records + self.page_size - 1) // self.page_size)
        
        # 更新分页控件状态
        self.prev_btn.setEnabled(self.current_page > 1)
        self.next_btn.setEnabled(self.current_page < total_pages)
        self.page_label.setText(
            f"第 {self.current_page}/{total_pages} 页 (共 {total_records} 条记录)"
        )        
        # 获取当前页数据
        offset = (self.current_page - 1) * self.page_size
        query = f'''
            SELECT id, name, batch, production_date, expiry_date, quantity, gtin
            FROM reagents 
            WHERE 1=1 AND quantity > 0
        '''
        
        if name_filter:
            query += " AND name LIKE ?"
        
        if batch_filter:
            query += " AND batch LIKE ?"
        
        query += " ORDER BY id LIMIT ? OFFSET ?"
        
        params = []
        if name_filter:
            params.append(f"%{name_filter}%")
        if batch_filter:
            params.append(f"%{batch_filter}%")
        
        params.extend([self.page_size, offset])
        
        self.db.cursor.execute(query, params)
        reagents = self.db.cursor.fetchall()
        
        # 获取当前日期
        current_date = QDate.currentDate()

        # 更新表格数据
        self.model.setRowCount(0)
        for row in reagents:
            # 解析有效期
            expiry_date = QDate.fromString(row[4], "yyyy-MM-dd")
            
            # 计算剩余天数
            days_remaining = current_date.daysTo(expiry_date) if expiry_date.isValid() else -1
            
            # 创建表格项
            row_items = [
                QStandardItem(str(row[0])),  # ID
                QStandardItem(str(row[1])),  # 试剂名称
                QStandardItem(str(row[2])),  # 批号
                QStandardItem(str(row[3])),  # 生产日期
                QStandardItem(str(row[4])),  # 有效期至
                QStandardItem(str(row[5])),  # 库存数量
                QStandardItem(str(days_remaining) if days_remaining >= 0 else "已过期"),  # 剩余天数
                QStandardItem(str(row[6]) if row[6] else "")  # GTIN
            ]
            
            # 设置日期列对齐方式
            for col in [3, 4]:  # 生产日期和有效期列
                row_items[col].setTextAlignment(Qt.AlignCenter)
            
            # 设置库存列对齐方式
            row_items[5].setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
            
            # 设置剩余天数列对齐方式
            row_items[6].setTextAlignment(Qt.AlignCenter)
            
            # 检查有效期状态
            if days_remaining < 0:  # 已过期
                for i in range(len(row_items)):
                    if i != 6:  # 除了剩余天数列
                        row_items[i].setForeground(QBrush(QColor(255, 0, 0)))  # 红色
                row_items[6].setText("已过期")
                row_items[6].setForeground(QBrush(QColor(255, 0, 0)))  # 红色
            elif days_remaining <= 30:  # 30天内过期
                row_items[6].setForeground(QBrush(QColor(255, 165, 0)))  # 橙色
            elif days_remaining <= 90:  # 90天内过期
                row_items[6].setForeground(QBrush(QColor(255, 215, 0)))  # 黄色
            else:  # 有效
                row_items[6].setForeground(QBrush(QColor(0, 128, 0)))  # 绿色
            
            self.model.appendRow(row_items)
        
        # 自动按ID排序（升序）
        self.table_view.sortByColumn(0, Qt.AscendingOrder)
    
    def prev_page(self):
        if self.current_page > 1:
            self.current_page -= 1
            self.load_data()
    
    def next_page(self):
        self.db.cursor.execute("SELECT COUNT(*) FROM reagents")
        total_records = self.db.cursor.fetchone()[0]
        total_pages = max(1, (total_records + self.page_size - 1) // self.page_size)
        
        if self.current_page < total_pages:
            self.current_page += 1
            self.load_data()
    
    def show_inbound_dialog(self):
        dialog = InboundDialog(self.db, self)
        if dialog.exec_() == QDialog.Accepted:
            self.load_data()
            # 刷新出入库记录页面
            self.combined_record_tab.load_data()
    
    def show_outbound_dialog(self):
        dialog = OutboundDialog(self.db, self)
        if dialog.exec_() == QDialog.Accepted:
            self.load_data()
            # 刷新出入库记录页面
            self.combined_record_tab.load_data()
    
    def show_name_manager(self):
        dialog = ReagentNameManager(self.db, self)
        dialog.exec_()

    def backup_database_on_uninstall():
        """在卸载时备份数据库文件"""
        try:
            # 确定数据库文件路径
            appdata_path = os.getenv('APPDATA')
            if not appdata_path:
                print("无法获取APPDATA环境变量")
                return False
            
            db_dir = os.path.join(appdata_path, "ReagentManagementSystem")
            db_file = os.path.join(db_dir, "ReagentWarehouseData.db")
            
            if not os.path.exists(db_file):
                print("数据库文件不存在，无需备份")
                return True
            
            # 创建备份目录
            backup_dir = os.path.join(db_dir, "Backups")
            os.makedirs(backup_dir, exist_ok=True)
            
            # 生成带时间戳的备份文件名
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_file = os.path.join(backup_dir, f"ReagentWarehouseData_{timestamp}.db")
            
            # 备份数据库文件
            shutil.copyfile(db_file, backup_file)
            print(f"数据库已备份到: {backup_file}")
            
            # 删除原数据库文件
            os.remove(db_file)
            print("原数据库文件已删除")
            
            # 如果数据库目录为空，则删除目录
            if not os.listdir(db_dir):
                os.rmdir(db_dir)
                print("数据库目录已删除")
            
            return True
        except Exception as e:
            print(f"备份过程中出错: {str(e)}")
            return False
        

    def cleanup_old_installer(self):
        """清理旧的安装包文件"""
        file = os.path.join(DB_DIR, "saved_path.txt")
        if not os.path.exists(file):
            return None
        
        with open(file, 'r') as f:
            file_path = f.read()

        if os.path.exists(file_path):
            try:
                os.remove(file_path)
                print(f"已删除安装包: {file_path}")
            except Exception as e:
                print(f"删除安装包失败: {str(e)}")

if __name__ == "__main__":
    config_path = os.path.join(application_path, 'config.ini')
    if not getattr(sys, 'frozen', False) and not os.path.exists(config_path):
        config = configparser.ConfigParser()
        config['Application'] = {
            'Version': '1.0.0'
        }
        config['Update'] = {
            'Server': 'http://127.0.0.1:8000'
        }
        config['Security'] = {
            
        }
        with open(config_path, 'w') as configfile:
            config.write(configfile)

    app = QApplication(sys.argv)
    
    # 设置应用样式
    app.setStyle("Fusion")
    app.setStyleSheet("""
        /* 全局样式 */
        Widget {
            font-family: 'Segoe UI', 'Microsoft YaHei', sans-serif;
            font-size: 10pt;
            background-color: #f5f7fa;
        

        MainWindow {
            background-color: #f0f4f8;
        

        * 按钮样式 */
        PushButton {
            background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                            stop:0 #4CAF50, stop:1 #2E7D32);
            color: white;
            border: none;
            padding: 8px 16px;
            border-radius: 4px;
            font-weight: bold;
            min-height: 30px;
        

        PushButton:hover {
            background-color: qlineargradient(x1:0, y1:0, x2:0, y2:1,
                                         stop:0 #66BB6A, stop:1 #388E3C);
        

        PushButton:pressed {
            background-color: #1B5E20;
        

        PushButton:disabled {
            background-color: #BDBDBD;
            color: #757575;
        

        * 表格样式 */
        TableView {
            gridline-color: #e0e0e0;
            alternate-background-color: #f8f9fa;
            background-color: white;
            font-size: 10pt;
            border: 1px solid #e0e0e0;
            border-radius: 4px;
        

        HeaderView::section {
            background-color: #e3f2fd;
            padding: 6px;
            border: none;
            font-weight: bold;
            font-size: 10pt;
            border-bottom: 2px solid #bbdefb;
        

        * 输入控件 */
        LineEdit, QComboBox, QDateEdit {
            padding: 6px;
            border: 1px solid #bdbdbd;
            border-radius: 4px;
            background-color: white;
            min-height: 28px;
        

        LineEdit:focus, QComboBox:focus, QDateEdit:focus {
            border: 1px solid #4fc3f7;
            background-color: #e1f5fe;
        

        * 对话框 */
        Dialog {
            background-color: #f5f7fa;
            border: 1px solid #90a4ae;
            border-radius: 6px;
        

        * 标签 */
        Label {
            font-weight: 500;
            color: #37474f;
        

        * 状态栏 */
        StatusBar {
            background-color: #e3f2fd;
            color: #0277bd;
            font-size: 9pt;
            border-top: 1px solid #bbdefb;
        

        * 选项卡 */
        TabWidget::pane {
            border: 1px solid #cfd8dc;
            background: white;
            border-radius: 4px;
            margin-top: 10px;
        

        TabBar::tab {
            background: #e3f2fd;
            border: 1px solid #bbdefb;
            border-bottom: none;
            padding: 8px 16px;
            margin-right: 2px;
            border-top-left-radius: 4px;
            border-top-right-radius: 4px;
            color: #01579b;
            font-weight: bold;
        

        TabBar::tab:selected {
            background: white;
            border-bottom: 2px solid #0288d1;
            color: #0288d1;
        

        TabBar::tab:hover {
           background: #bbdefb;

        QPushButton {
            transition: background-image 0.3s ease;
        }

        * 列表控件 */
        ListWidget {
           border: 1px solid #bdbdbd;
           background: white;
           font-size: 10pt;
           border-radius: 4px;
        

        ListWidget::item {
           padding: 8px;
           border-bottom: 1px solid #e0e0e0;
        

        ListWidget::item:selected {
           background-color: #4fc3f7;
           color: white;
        

        * 分组框 */
        GroupBox {
           border: 1px solid #cfd8dc;
           border-radius: 4px;
           margin-top: 1ex;
           font-weight: bold;
           background-color: #e3f2fd;
           padding: 10px;
        

        GroupBox::title {
           subcontrol-origin: margin;
           subcontrol-position: top left;
           padding: 0 5px;
           background-color: transparent;
        

        * 工具提示 */
        ToolTip {
           background-color: #fffeeb;
           color: #37474f;
           border: 1px solid #ffd54f;
           border-radius: 4px;
           padding: 5px;
        
    """)

    # test1 = f"0106936415919665{chr(29)}11240819{chr(29)}17260218{chr(29)}102024050123{chr(29)}97105-002512-00{chr(29)}98AAX202405012303531"
    # parsed = GS1Parser.parse_barcode(test1)
    # print("1",parsed)

    window = ReagentManagementSystem()
    window.show()
    sys.exit(app.exec_())